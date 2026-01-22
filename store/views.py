from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_protect
from datetime import timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import lru_cache
import re

from .payments import razorpay as razorpay_provider

def is_customer(user):
    return user.is_authenticated and user.role == 'customer'

def is_staff(user):
    return user.is_authenticated and (user.role == 'staff' or user.role == 'admin')

def is_admin(user):
    return user.is_authenticated and user.role == 'admin'


GUEST_CART_SESSION_KEY = 'guest_cart'


def _ensure_guest_cart(request):
    cart = request.session.get(GUEST_CART_SESSION_KEY)
    if not isinstance(cart, dict):
        cart = {}
    cart.setdefault('fish', {})
    cart.setdefault('accessories', {})
    cart.setdefault('plants', {})
    request.session[GUEST_CART_SESSION_KEY] = cart
    return cart


def _save_guest_cart(request, cart):
    request.session[GUEST_CART_SESSION_KEY] = cart
    request.session.modified = True


def _guest_cart_total_items(cart):
    total = 0
    for section in ('fish', 'accessories', 'plants'):
        entries = cart.get(section, {})
        if isinstance(entries, dict):
            for data in entries.values():
                try:
                    total += int(data.get('quantity', 0))
                except (TypeError, ValueError):
                    continue
    return total


def _guest_cart_key(item_id, combo_id=None):
    return f"{item_id}:{combo_id if combo_id else 'none'}"


def _add_guest_fish(request, fish_id, quantity, combo_id=None):
    cart = _ensure_guest_cart(request)
    key = _guest_cart_key(fish_id, combo_id)
    entry = cart['fish'].get(key, {'fish_id': fish_id, 'quantity': 0, 'combo_id': combo_id})
    entry['quantity'] = int(entry.get('quantity', 0) or 0) + int(quantity or 0)
    cart['fish'][key] = entry
    _save_guest_cart(request, cart)
    return cart


def _add_guest_accessory(request, accessory_id, quantity):
    cart = _ensure_guest_cart(request)
    key = str(accessory_id)
    entry = cart['accessories'].get(key, {'accessory_id': accessory_id, 'quantity': 0})
    entry['quantity'] = int(entry.get('quantity', 0) or 0) + int(quantity or 0)
    cart['accessories'][key] = entry
    _save_guest_cart(request, cart)
    return cart


def _add_guest_plant(request, plant_id, quantity):
    cart = _ensure_guest_cart(request)
    key = str(plant_id)
    entry = cart['plants'].get(key, {'plant_id': plant_id, 'quantity': 0})
    entry['quantity'] = int(entry.get('quantity', 0) or 0) + int(quantity or 0)
    cart['plants'][key] = entry
    _save_guest_cart(request, cart)
    return cart


def _guest_cart_total_counter(request):
    cart = _ensure_guest_cart(request)
    return _guest_cart_total_items(cart)


def _merge_guest_cart_into_user(request, user):
    if not user or getattr(user, 'role', None) != 'customer':
        # Non-customer logins should not keep guest carts; clear to avoid reuse.
        if GUEST_CART_SESSION_KEY in request.session:
            request.session.pop(GUEST_CART_SESSION_KEY, None)
            request.session.modified = True
        return

    cart = request.session.get(GUEST_CART_SESSION_KEY)
    if not isinstance(cart, dict):
        return

    fish_entries = cart.get('fish', {}) if isinstance(cart.get('fish'), dict) else {}
    accessory_entries = cart.get('accessories', {}) if isinstance(cart.get('accessories'), dict) else {}
    plant_entries = cart.get('plants', {}) if isinstance(cart.get('plants'), dict) else {}

    try:
        for data in fish_entries.values():
            fish_id = data.get('fish_id')
            quantity = max(int(data.get('quantity', 0) or 0), 0)
            if not fish_id or quantity <= 0:
                continue
            fish = Fish.objects.filter(id=fish_id).first()
            if not fish:
                continue
            combo_obj = None
            combo_id = data.get('combo_id')
            if combo_id:
                combo_obj = ComboOffer.objects.filter(id=combo_id).first()
            cart_item, created = Cart.objects.get_or_create(
                user=user,
                fish=fish,
                combo=combo_obj,
                defaults={'quantity': quantity}
            )
            if not created:
                cart_item.quantity += quantity
                cart_item.save()

        for data in accessory_entries.values():
            accessory_id = data.get('accessory_id')
            quantity = max(int(data.get('quantity', 0) or 0), 0)
            if not accessory_id or quantity <= 0:
                continue
            accessory = Accessory.objects.filter(id=accessory_id).first()
            if not accessory:
                continue
            acc_item, created = AccessoryCart.objects.get_or_create(
                user=user,
                accessory=accessory,
                defaults={'quantity': quantity}
            )
            if not created:
                acc_item.quantity += quantity
                acc_item.save()

        for data in plant_entries.values():
            plant_id = data.get('plant_id')
            quantity = max(int(data.get('quantity', 0) or 0), 0)
            if not plant_id or quantity <= 0:
                continue
            plant = Plant.objects.filter(id=plant_id).first()
            if not plant or plant.price is None:
                continue
            plant_item, created = PlantCart.objects.get_or_create(
                user=user,
                plant=plant,
                defaults={'quantity': quantity}
            )
            if not created:
                plant_item.quantity += quantity
                plant_item.save()
    finally:
        request.session.pop(GUEST_CART_SESSION_KEY, None)
        request.session.modified = True


class GuestFishCartItem:
    def __init__(self, fish, quantity, combo=None):
        self.fish = fish
        self.quantity = max(int(quantity or 0), 0)
        self.combo = combo
        self.combo_id = getattr(combo, 'id', None)
        # expose fish_id so guest items can share combo logic with DB-backed items
        self.fish_id = getattr(fish, 'id', None)

    def get_total(self):
        try:
            return Decimal(self.fish.price) * Decimal(self.quantity)
        except Exception:
            return Decimal('0')


class GuestAccessoryCartItem:
    def __init__(self, accessory, quantity):
        self.accessory = accessory
        self.quantity = max(int(quantity or 0), 0)

    def get_total(self):
        try:
            return Decimal(self.accessory.price) * Decimal(self.quantity)
        except Exception:
            return Decimal('0')


class GuestPlantCartItem:
    def __init__(self, plant, quantity):
        self.plant = plant
        self.quantity = max(int(quantity or 0), 0)

    def get_total(self):
        try:
            return Decimal(self.plant.price) * Decimal(self.quantity)
        except Exception:
            return Decimal('0')


def _build_guest_cart_items(request):
    cart = _ensure_guest_cart(request)

    fish_entries = cart.get('fish', {})
    accessory_entries = cart.get('accessories', {})
    plant_entries = cart.get('plants', {})

    fish_ids = {int(data.get('fish_id')) for data in fish_entries.values() if data.get('fish_id')}
    accessory_ids = {int(data.get('accessory_id')) for data in accessory_entries.values() if data.get('accessory_id')}
    plant_ids = {int(data.get('plant_id')) for data in plant_entries.values() if data.get('plant_id')}
    combo_ids = {int(data.get('combo_id')) for data in fish_entries.values() if data.get('combo_id')}

    fish_map = {f.id: f for f in Fish.objects.filter(id__in=fish_ids)} if fish_ids else {}
    accessory_map = {a.id: a for a in Accessory.objects.filter(id__in=accessory_ids)} if accessory_ids else {}
    plant_map = {p.id: p for p in Plant.objects.filter(id__in=plant_ids)} if plant_ids else {}
    combo_map = {c.id: c for c in ComboOffer.objects.filter(id__in=combo_ids)} if combo_ids else {}

    guest_fish_items = []
    for key, data in list(fish_entries.items()):
        fish_id = data.get('fish_id')
        if not fish_id or fish_id not in fish_map:
            fish_entries.pop(key, None)
            continue
        quantity = max(int(data.get('quantity', 0) or 0), 0)
        if quantity <= 0:
            fish_entries.pop(key, None)
            continue
        combo_id = data.get('combo_id')
        combo = combo_map.get(int(combo_id)) if combo_id else None
        guest_fish_items.append(GuestFishCartItem(fish_map[fish_id], quantity, combo))

    guest_accessories = []
    for key, data in list(accessory_entries.items()):
        accessory_id = data.get('accessory_id')
        if not accessory_id or accessory_id not in accessory_map:
            accessory_entries.pop(key, None)
            continue
        quantity = max(int(data.get('quantity', 0) or 0), 0)
        if quantity <= 0:
            accessory_entries.pop(key, None)
            continue
        guest_accessories.append(GuestAccessoryCartItem(accessory_map[accessory_id], quantity))

    guest_plants = []
    for key, data in list(plant_entries.items()):
        plant_id = data.get('plant_id')
        if not plant_id or plant_id not in plant_map:
            plant_entries.pop(key, None)
            continue
        quantity = max(int(data.get('quantity', 0) or 0), 0)
        if quantity <= 0:
            plant_entries.pop(key, None)
            continue
        guest_plants.append(GuestPlantCartItem(plant_map[plant_id], quantity))

    _save_guest_cart(request, cart)
    return guest_fish_items, guest_accessories, guest_plants


KERALA_PIN_PREFIXES = ('67', '68', '69')


def _calculate_total_weight(cart_items, accessory_items, plant_items):
    """Compute total cart weight in kilograms, respecting combo bundle weights."""
    total_weight = Decimal('0')

    combo_groups = {}
    combo_ids = set()
    standalone_items = []

    for item in cart_items:
        if item.combo_id:
            combo_groups.setdefault(item.combo_id, {'items': []})['items'].append(item)
            combo_ids.add(item.combo_id)
        else:
            fish_weight = Decimal(item.fish.weight or 0)
            total_weight += fish_weight * Decimal(item.quantity or 0)

    combos_prefetched = {}
    if combo_ids:
        combos_prefetched = {
            combo.id: combo
            for combo in ComboOffer.objects.filter(id__in=combo_ids).prefetch_related('items__fish')
        }

    for combo_id, data in combo_groups.items():
        items = data['items']
        combo = combos_prefetched.get(combo_id)
        if not combo:
            for it in items:
                fish_weight = Decimal(it.fish.weight or 0)
                total_weight += fish_weight * Decimal(it.quantity or 0)
            continue

        combo_items = list(combo.items.all())
        requirements = {ci.fish_id: int(ci.quantity or 1) for ci in combo_items}
        bundle_counts = []
        for it in items:
            req = max(1, requirements.get(it.fish_id, 1))
            try:
                bundle_counts.append(int(it.quantity) // req)
            except Exception:
                bundle_counts.append(0)
        bundle_count = min(bundle_counts) if bundle_counts else 0

        if combo.weight is not None and bundle_count > 0:
            total_weight += Decimal(combo.weight) * Decimal(bundle_count)
        elif bundle_count > 0:
            for ci in combo_items:
                fish_weight = Decimal(ci.fish.weight or 0)
                total_weight += fish_weight * Decimal(int(ci.quantity or 1) * bundle_count)

        for it in items:
            req = max(1, requirements.get(it.fish_id, 1)) if requirements else 1
            assigned = bundle_count * req
            leftover = max(0, int(it.quantity) - assigned)
            if leftover:
                fish_weight = Decimal(it.fish.weight or 0)
                total_weight += fish_weight * Decimal(leftover)

    for accessory_item in accessory_items:
        accessory_weight = Decimal(accessory_item.accessory.weight or 0)
        total_weight += accessory_weight * Decimal(accessory_item.quantity or 0)

    for plant_item in plant_items:
        plant_weight = Decimal(plant_item.plant.weight or 0)
        total_weight += plant_weight * Decimal(plant_item.quantity or 0)

    return total_weight


def _is_kerala_destination(state=None, pincode=None, address=None):
    state_value = (state or '').strip().lower()
    digits = ''.join(ch for ch in (pincode or '') if ch.isdigit())

    state_matches = state_value == 'kerala'
    pin_matches = len(digits) >= 2 and digits[:2] in KERALA_PIN_PREFIXES

    # When both state and pincode are supplied, require them to agree before
    # granting Kerala shipping rates to avoid mismatched combinations.
    if state_value and digits:
        return state_matches and pin_matches

    # Fall back to whichever signal is available if only one field is present.
    if state_matches or pin_matches:
        return True

    if address and 'kerala' in address.lower():
        return True

    return False


STATE_SPLIT_PATTERN = re.compile(r'[\n,]+')


def _normalize_state_value(value):
    if not value:
        return ''
    return re.sub(r'\s+', ' ', value).strip().lower()


def _parse_unserviceable_states(raw_value):
    if not raw_value:
        return []
    cleaned = raw_value.replace('\r', '\n')
    entries = [part.strip() for part in STATE_SPLIT_PATTERN.split(cleaned) if part.strip()]
    return entries


class ShippingUnavailableError(Exception):
    def __init__(self, state_display):
        state_display = (state_display or '').strip() or 'the selected state'
        self.state = state_display
        super().__init__(f'Delivery is not available in {state_display}.')


@lru_cache(maxsize=1)
def _get_shipping_rates():
    try:
        setting, _ = ShippingChargeSetting.objects.get_or_create(
            key='default',
            defaults={'kerala_rate': Decimal('60.00'), 'default_rate': Decimal('100.00')},
        )
        kerala_rate = Decimal(setting.kerala_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        default_rate = Decimal(setting.default_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        location_rates = {
            _normalize_state_value(entry.location_name): Decimal(entry.shipping_charge).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            for entry in ShippingChargeByLocation.objects.all()
            if _normalize_state_value(entry.location_name)
        }
        unserviceable = tuple(
            _normalize_state_value(entry)
            for entry in _parse_unserviceable_states(getattr(setting, 'unserviceable_states', ''))
            if _normalize_state_value(entry)
        )
        return kerala_rate, default_rate, unserviceable, location_rates
    except Exception:
        logging.getLogger(__name__).exception('Falling back to default shipping rates')
        return Decimal('60.00'), Decimal('100.00'), tuple(), {}


def _calculate_delivery_charge(total_weight, state=None, pincode=None, address=None):
    kerala_rate, default_rate, unserviceable_states, location_rates = _get_shipping_rates()
    normalized_state = _normalize_state_value(state)
    if normalized_state and normalized_state in unserviceable_states:
        raise ShippingUnavailableError(state)

    if normalized_state and normalized_state in location_rates:
        rate = location_rates[normalized_state]
    else:
        is_kerala = _is_kerala_destination(state=state, pincode=pincode, address=address)
        rate = kerala_rate if is_kerala else default_rate

    try:
        weight = Decimal(total_weight)
    except (TypeError, InvalidOperation):
        weight = Decimal('0')

    if weight <= 0:
        return Decimal('0.00'), rate

    billable_weight = max(weight, Decimal('1.00'))
    charge = (billable_weight * rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return charge, rate

@login_required
@user_passes_test(is_customer)
def checkout_view(request):
    """Render the checkout page with cart, coupons and payment options."""
    cart_items = Cart.objects.filter(user=request.user).select_related('fish', 'combo')
    accessory_items = AccessoryCart.objects.filter(user=request.user).select_related('accessory')
    plant_items = PlantCart.objects.filter(user=request.user).select_related('plant')

    if not cart_items and not accessory_items and not plant_items:
        messages.error(request, 'Your cart is empty.')
        return redirect('cart')

    # Group cart items by combo
    bundle_groups = []
    standalone_items = []
    combos_map = {}
    for item in cart_items:
        if item.combo_id:
            combos_map.setdefault(item.combo_id, {'combo': item.combo, 'items': []})['items'].append(item)
        else:
            standalone_items.append(item)

    for _, group in combos_map.items():
        combo = group.get('combo')
        bundle_count = 0
        per_bundle_value = Decimal('0')
        if combo:
            combo_items = {ci.fish_id: ci.quantity for ci in combo.items.all()}
            per_bundle_value = sum(Decimal(ci.fish.price) * Decimal(int(ci.quantity)) for ci in combo.items.select_related('fish').all())
            counts = []
            for it in group.get('items', []):
                req = combo_items.get(it.fish_id, 1)
                try:
                    counts.append(int(it.quantity) // int(req))
                except Exception:
                    counts.append(0)
            bundle_count = min(counts) if counts else 0

        if combo and combo.bundle_price:
            group['display_price'] = Decimal(combo.bundle_price) * Decimal(bundle_count)
        else:
            group['display_price'] = Decimal(bundle_count) * per_bundle_value
        group['bundle_count'] = bundle_count
        bundle_groups.append(group)

    # Compute total considering bundle pricing when applicable
    total = Decimal('0')
    total += sum(Decimal(item.get_total()) for item in standalone_items)
    for g in bundle_groups:
        combo = g.get('combo')
        items = g.get('items', [])
        bundle_count = g.get('bundle_count', 0)
        if combo and combo.bundle_price and bundle_count > 0:
            total += Decimal(combo.bundle_price) * Decimal(bundle_count)
            combo_items = {ci.fish_id: ci.quantity for ci in combo.items.all()}
            for it in items:
                req = combo_items.get(it.fish_id, 1)
                leftover = int(it.quantity) - (int(req) * bundle_count)
                if leftover > 0:
                    total += Decimal(it.fish.price) * Decimal(leftover)
        else:
            total += sum(Decimal(it.get_total()) for it in items)

    total += sum(Decimal(item.get_total()) for item in accessory_items)
    total += sum(Decimal(item.get_total()) for item in plant_items)

    total_weight = _calculate_total_weight(cart_items, accessory_items, plant_items)

    # Applied coupon from session
    applied_coupon = None
    discount = Decimal('0')
    final_total = total
    if 'applied_coupon_code' in request.session:
        try:
            coupon = Coupon.objects.get(code=request.session['applied_coupon_code'])
            if coupon.is_valid() and coupon.can_use(request.user):
                applied_coupon = coupon
                discount = (total * Decimal(coupon.discount_percentage)) / Decimal('100')
                if coupon.max_discount_amount:
                    discount = min(discount, Decimal(coupon.max_discount_amount))
                final_total = total - discount
        except Coupon.DoesNotExist:
            request.session.pop('applied_coupon_code', None)

    # Prefill shipping details from last order or user profile
    last_order = Order.objects.filter(user=request.user).order_by('-created_at').first()
    shipping_address_prefill = ''
    shipping_state_prefill = ''
    shipping_pincode_prefill = ''
    if last_order:
        shipping_address_prefill = last_order.shipping_address or ''
        shipping_state_prefill = last_order.shipping_state or ''
        shipping_pincode_prefill = last_order.shipping_pincode or ''
    if not shipping_address_prefill:
        shipping_address_prefill = getattr(request.user, 'address', '') or ''

    shipping_setting = ShippingChargeSetting.objects.filter(key='default').only('unserviceable_states').first()
    unserviceable_state_list = _parse_unserviceable_states(
        getattr(shipping_setting, 'unserviceable_states', '') if shipping_setting else ''
    )

    shipping_blocked_state = None
    try:
        delivery_charge, delivery_rate = _calculate_delivery_charge(
            total_weight,
            state=shipping_state_prefill,
            pincode=shipping_pincode_prefill,
            address=shipping_address_prefill,
        )
    except ShippingUnavailableError as exc:
        shipping_blocked_state = exc.state
        delivery_charge = Decimal('0.00')
        delivery_rate = Decimal('0.00')
        messages.error(
            request,
            f'Delivery is not available in {exc.state}. Please choose a different state to continue.',
        )

    net_total = max(Decimal('0'), final_total)
    final_total = net_total + delivery_charge

    kerala_delivery_rate, default_delivery_rate, _, location_delivery_rates = _get_shipping_rates()
    location_delivery_rates = {
        key: float(rate) for key, rate in (location_delivery_rates or {}).items()
    }

    # Available coupon suggestions (simple rules)
    now = timezone.now()
    coupon_types_allowed = ['all']
    if getattr(request.user, 'is_favorite', False):
        coupon_types_allowed.append('favorites')
    else:
        coupon_types_allowed.append('normal')

    available_coupons = Coupon.objects.filter(
        is_active=True,
        show_in_suggestions=True,
        valid_from__lte=now,
        valid_until__gte=now,
        coupon_type__in=coupon_types_allowed,
    ).exclude(
        usage_limit__isnull=False,
        times_used__gte=models.F('usage_limit')
    ).filter(
        Q(min_order_amount__isnull=True) | Q(min_order_amount__lte=total)
    ).order_by('-discount_percentage')[:5]

    return render(request, 'store/customer/checkout.html', {
        'cart_items': cart_items,
        'bundle_groups': bundle_groups,
        'standalone_items': standalone_items,
        'accessory_items': accessory_items,
        'plant_items': plant_items,
        'total': total,
        'applied_coupon': applied_coupon,
        'discount': discount,
        'final_total': final_total,
        'available_coupons': available_coupons,
        'total_weight': total_weight,
        'delivery_charge': delivery_charge,
        'delivery_rate': delivery_rate,
        'kerala_delivery_rate': kerala_delivery_rate,
        'default_delivery_rate': default_delivery_rate,
        'location_delivery_rates': location_delivery_rates,
        'shipping_address_prefill': shipping_address_prefill,
        'shipping_state_prefill': shipping_state_prefill,
        'shipping_pincode_prefill': shipping_pincode_prefill,
        'unserviceable_state_names': unserviceable_state_list,
        'shipping_blocked_state': shipping_blocked_state,
    })
from .models import CustomUser, Category, Breed, Fish, Order, OrderItem, Review, Service, ContactInfo, Coupon, LimitedOffer
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.mail import send_mail, EmailMultiAlternatives
from django.http import JsonResponse, HttpResponse
from django.db import models, transaction
from django.db.models import Q, Sum, Count
from django.conf import settings
from django.urls import reverse
from django.contrib.auth import update_session_auth_hash
from django.utils import timezone
from django.template.loader import render_to_string
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import (
    CustomUser,
    Category,
    Breed,
    Fish,
    FishMedia,
    Cart,
    AccessoryCart,
    Order,
    OrderItem,
    OrderAccessoryItem,
    OrderPlantItem,
    OTP,
    Review,
    Service,
    ContactInfo,
    Coupon,
    Accessory,
    ContactGalleryMedia,
    Plant,
    PlantCart,
    PlantMedia,
    ShippingChargeSetting,
    ShippingChargeByLocation,
)
from django.contrib import messages
from .forms import (
    CustomUserCreationForm,
    StaffCreateForm,
    CategoryForm,
    BreedForm,
    FishForm,
    FishMediaForm,
    ProfileEditForm,
    OrderFilterForm,
    ChangePasswordForm,
    ReviewForm,
    ServiceForm,
    ContactInfoForm,
    CouponForm,
    LimitedOfferForm,
    PlantForm,
    PlantMediaForm,
    ShippingChargeForm,
    ShippingChargeByLocationForm,
)
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import user_passes_test
from django import forms
from .models import ComboOffer, ComboItem
from urllib.parse import quote_plus
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import qrcode
from io import BytesIO
import base64
import uuid
import smtplib
import socket
import logging
from collections import defaultdict
from django.contrib.sessions.models import Session
import os


def is_customer(user):
    return user.is_authenticated and user.role == 'customer'

def is_staff_user(user):
    return user.is_authenticated and (user.role == 'staff' or user.is_superuser)


def is_staff(user):
    return user.is_authenticated and (user.role == 'staff' or user.role == 'admin')


def is_admin(user):
    return user.is_authenticated and user.role == 'admin'


def combos_view(request):
    """Public page: list active combos for customers."""
    # Prefetch related fishes to avoid N+1 queries
    all_combos = ComboOffer.objects.filter(is_active=True).prefetch_related('items__fish').order_by('-created_at')

    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '').strip()
    combo_categories = Category.objects.filter(category_type='combo').order_by('name')

    visible_combos = []
    for combo in all_combos:
        ok = True
        for item in combo.items.all():
            fish = item.fish
            # Require fish to be available and have at least the required quantity
            req_qty = int(item.quantity or 1)
            if not getattr(fish, 'is_available', True) or getattr(fish, 'stock_quantity', 0) < req_qty:
                ok = False
                break
        if ok:
            # prepare up-to-4 preview items for collage (pad with None for placeholders)
            items_list = list(combo.items.all()[:4])
            if len(items_list) < 4:
                items_list = items_list + [None] * (4 - len(items_list))
            combo.preview_items = items_list
            visible_combos.append(combo)

    # Apply search filter by combo title or included fish names
    if search_query:
        sq = search_query.lower()
        filtered = []
        for combo in visible_combos:
            title_match = sq in (combo.title or '').lower()
            fish_match = False
            if not title_match:
                for item in combo.items.all():
                    name = getattr(item.fish, 'name', '')
                    if name and sq in name.lower():
                        fish_match = True
                        break
            if title_match or fish_match:
                filtered.append(combo)
        visible_combos = filtered

    # Filter by combo category if requested
    if category_filter:
        try:
            cat_id = int(category_filter)
        except ValueError:
            cat_id = None
        if cat_id:
            visible_combos = [combo for combo in visible_combos if getattr(combo.category, 'id', None) == cat_id]

    # If AJAX request, return partial rendering for live updates
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render(request, 'store/customer/partials/_combo_cards.html', {'combos': visible_combos}).content.decode('utf-8')
        return JsonResponse({'html': html})

    return render(request, 'store/customer/combos.html', {
        'combos': visible_combos,
        'search_query': search_query,
        'category_filter': category_filter,
        'combo_categories': combo_categories,
    })


def combo_detail_view(request, combo_id):
    """Public combo detail page showing included fishes and bundle price."""
    combo = get_object_or_404(ComboOffer, id=combo_id, is_active=True)
    items = combo.items.select_related('fish').all()
    # Prepare preview items (up to 4) for collage display
    preview_items = list(items[:4])
    if len(preview_items) < 4:
        preview_items = preview_items + [None] * (4 - len(preview_items))

    return render(request, 'store/customer/combo_detail.html', {
        'combo': combo,
        'items': items,
        'preview_items': preview_items,
    })


@login_required
@user_passes_test(is_staff)
def notifications_center_view(request):
    """Staff-only notifications center showing recent notifications."""
    from .models import Notification
    qs = Notification.objects.all().order_by('-created_at')
    return render(request, 'store/notifications.html', {'notifications': qs})


@login_required
@user_passes_test(is_staff)
@require_POST
def mark_notifications_read_view(request):
    """Mark notifications as read. Accepts POST.

    - If `all=1` in POST, mark all unread as read.
    - Else if `ids[]` provided, mark those ids as read.
    - Else if `id` provided, mark that one as read.
    Returns JSON with number marked.
    """
    from .models import Notification
    marked = 0
    try:
        if request.POST.get('all') == '1':
            marked = Notification.objects.filter(is_read=False).update(is_read=True)
        else:
            ids = request.POST.getlist('ids[]') or ([request.POST.get('id')] if request.POST.get('id') else [])
            ids = [int(x) for x in ids if x]
            if ids:
                marked = Notification.objects.filter(id__in=ids, is_read=False).update(is_read=True)
    except Exception:
        return JsonResponse({'success': False, 'marked': 0})

    return JsonResponse({'success': True, 'marked': int(marked)})


# Authentication Views
def register_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            # Do not create user yet; store form data in session
            data = {
                'username': form.cleaned_data.get('username'),
                'email': form.cleaned_data.get('email'),
                'password1': form.cleaned_data.get('password1'),
                'first_name': form.cleaned_data.get('first_name', ''),
                'last_name': form.cleaned_data.get('last_name', ''),
            }
            request.session['pending_registration'] = data

            # Generate OTP for email verification (session-based)
            otp_code = OTP.generate_otp()
            request.session['pending_registration_otp'] = otp_code
            request.session['pending_registration_time'] = timezone.now().isoformat()

            # Send OTP email
            send_mail(
                f'Email Verification OTP - {settings.SITE_NAME}',
                f"Your OTP for email verification is: {otp_code}\n\nThis OTP will expire in 5 minutes.",
                settings.DEFAULT_FROM_EMAIL or 'noreply@aquafishstore.com',
                [data['email']],
                fail_silently=False,
            )

            messages.success(request, 'We\'ve sent an OTP to your email. Please verify to complete registration.')
            return redirect('verify_otp')
    else:
        form = CustomUserCreationForm()

    return render(request, 'store/register.html', {'form': form})


def verify_otp_view(request):
    # Registration OTP verification using session
    pending = request.session.get('pending_registration')
    if not pending:
        messages.error(request, 'No registration in progress. Please register first.')
        return redirect('register')

    pending_email = pending.get('email')
    created_iso = request.session.get('pending_registration_time')
    try:
        created_at = timezone.datetime.fromisoformat(created_iso) if created_iso else None
        if created_at and timezone.is_naive(created_at):
            created_at = timezone.make_aware(created_at, timezone.get_current_timezone())
    except Exception:
        created_at = None

    if request.method == 'POST':
        otp_code = (request.POST.get('otp_code') or '').strip()
        session_otp = request.session.get('pending_registration_otp')

        # Expiry check: 5 minutes
        is_expired = False
        if created_at:
            delta = timezone.now() - created_at
            is_expired = delta.total_seconds() > 300

        if is_expired:
            messages.error(request, 'OTP has expired. Please request a new one.')
            return redirect('resend_otp')

        if session_otp and otp_code == session_otp:
            # Create user now
            username = pending.get('username')
            email = pending.get('email')
            password = pending.get('password1')
            first_name = pending.get('first_name', '')
            last_name = pending.get('last_name', '')

            if CustomUser.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists. Please register again with a different username.')
                # Clear pending session
                for k in ['pending_registration', 'pending_registration_otp', 'pending_registration_time']:
                    request.session.pop(k, None)
                return redirect('register')
            if CustomUser.objects.filter(email=email).exists():
                messages.error(request, 'Email already registered. Try logging in or reset your password.')
                for k in ['pending_registration', 'pending_registration_otp', 'pending_registration_time']:
                    request.session.pop(k, None)
                return redirect('login')

            user = CustomUser(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                role='customer',
                is_active=True,
                email_verified=True,
            )
            user.set_password(password)
            # Save with error handling to surface DB/validation issues
            logger = logging.getLogger(__name__)
            try:
                user.save()
            except Exception as exc:
                logger.exception('Failed to create user during OTP verification: %s', exc)
                messages.error(request, 'Unable to create account at this time. Please try again or contact support.')
                # Keep pending session so user can retry (do not pop yet)
                return redirect('verify_otp')

            # Clear pending session
            for k in ['pending_registration', 'pending_registration_otp', 'pending_registration_time']:
                request.session.pop(k, None)

            messages.success(request, 'Email verified and account created successfully! You can now login.')
            return redirect('login')
        else:
            messages.error(request, 'Invalid OTP. Please try again.')

    return render(request, 'store/verify_otp.html', {'pending_email': pending_email})


def terminate_user_sessions(user):
    """Delete all session records that belong to `user` so any active logins are ended.

    Returns the number of sessions removed.
    """
    removed = 0
    try:
        for session in Session.objects.all():
            try:
                data = session.get_decoded()
            except Exception:
                continue
            # Django stores user id in session key '_auth_user_id'
            if str(data.get('_auth_user_id')) == str(user.pk):
                session.delete()
                removed += 1
    except Exception:
        logging.getLogger(__name__).exception('Error terminating sessions for user %s', getattr(user, 'pk', None))
    return removed


def _send_order_email(order, template_base, subject, recipient_email, request=None):
    """Render email templates and send an order-related email (invoice/cancellation).

    `template_base` should be the base filename under `store/emails/`, e.g. 'invoice' or 'order_cancelled'.
    """
    try:
        context = {
            'order': order,
            'order_items': order.items.all(),
            'plant_items': order.plant_items.all() if hasattr(order, 'plant_items') else [],
            'accessory_items': order.accessory_items.all() if hasattr(order, 'accessory_items') else [],
            'site_name': settings.SITE_NAME,
        }
        if request is not None:
            try:
                context['order_url'] = request.build_absolute_uri(reverse('order_detail', args=[order.id]))
            except Exception:
                context['order_url'] = None

        text_body = render_to_string(f'store/emails/{template_base}.txt', context)
        html_body = render_to_string(f'store/emails/{template_base}.html', context)

        # Use EmailMultiAlternatives to support HTML alternative and attachments
        email = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=settings.DEFAULT_FROM_EMAIL or 'noreply@aquafishstore.com',
            to=[recipient_email],
        )
        email.extra_headers = {'Reply-To': settings.DEFAULT_FROM_EMAIL or 'noreply@aquafishstore.com'}
        # Attach HTML alternative
        email.attach_alternative(html_body, 'text/html')

        # If sending invoice, generate PDF and save to media (fallback). Attaching
        # the PDF to the email is controlled by `settings.INVOICE_ATTACHMENTS`.
        download_url = None
        pdf_bytes = None
        if template_base == 'invoice':
            try:
                pdf_bytes = generate_invoice_pdf(order)
                if pdf_bytes:
                    # Save a copy to MEDIA so we can include a download link if email delivery fails
                    try:
                        invoices_dir = os.path.join(getattr(settings, 'MEDIA_ROOT', os.path.join(getattr(settings, 'BASE_DIR', ''), 'media')), 'invoices')
                        os.makedirs(invoices_dir, exist_ok=True)
                        filename = f'invoice-{order.order_number}.pdf'
                        fs_path = os.path.join(invoices_dir, filename)
                        with open(fs_path, 'wb') as f:
                            f.write(pdf_bytes)
                        # Build URL: prefer absolute if request present
                        rel_url = os.path.join(getattr(settings, 'MEDIA_URL', '/media/'), 'invoices', filename).replace('\\', '/')
                        if request is not None:
                            try:
                                download_url = request.build_absolute_uri(rel_url)
                            except Exception:
                                download_url = rel_url
                        else:
                            download_url = rel_url
                    except Exception:
                        logging.getLogger(__name__).exception('Failed to save invoice PDF to media for order %s', order.order_number)
            except Exception:
                logging.getLogger(__name__).exception('Failed to generate PDF invoice for order %s', order.order_number)

        # Attach the generated PDF to the email when configured. We do not
        # include a download link in the email body to avoid exposing direct
        # media links in emails.
        try:
            if getattr(settings, 'INVOICE_ATTACHMENTS', False) and pdf_bytes:
                email.attach(f'invoice-{order.order_number}.pdf', pdf_bytes, 'application/pdf')
        except Exception:
            logging.getLogger(__name__).exception('Failed to attach PDF to email for order %s', order.order_number)

        try:
            email.send(fail_silently=False)
            logging.getLogger(__name__).info('Sent order email %s for order %s to %s', template_base, order.order_number, recipient_email)
        except Exception:
            # First send failed (often SMTP disconnects with attachments). Try again without attachments so customer still receives notification.
            logging.getLogger(__name__).exception('Failed to send order email %s for order %s; retrying without attachments', template_base, order.order_number)
            try:
                # Remove attachments and resend plain email (text + html alternative)
                email.attachments = []
                email.send(fail_silently=False)
                logging.getLogger(__name__).warning('Sent order email %s for order %s WITHOUT attachment to %s', template_base, order.order_number, recipient_email)
            except Exception:
                logging.getLogger(__name__).exception('Retry without attachment also failed for order %s', order.order_number)
    except Exception:
        logging.getLogger(__name__).exception('Failed to send order email %s for order %s', template_base, getattr(order, 'order_number', None))


def _ensure_order_inventory_deducted_locked(order):
    """Deduct fish/accessory stock for a locked Order instance."""
    if getattr(order, '_inventory_deducted', False):
        return False

    if getattr(order, 'status', None) != 'pending':
        return False

    logger = logging.getLogger(__name__)
    deducted = False

    fish_items = list(order.items.select_related('fish')) if hasattr(order, 'items') else []
    accessory_items = list(order.accessory_items.select_related('accessory')) if hasattr(order, 'accessory_items') else []
    plant_items = list(order.plant_items.select_related('plant')) if hasattr(order, 'plant_items') else []

    if fish_items:
        qty_by_fish = defaultdict(int)
        for item in fish_items:
            try:
                qty_by_fish[item.fish_id] += int(item.quantity or 0)
            except Exception:
                logger.exception('Invalid quantity for fish item %s on order %s', getattr(item, 'id', None), getattr(order, 'order_number', None))
        if qty_by_fish:
            fishes = {fish.id: fish for fish in Fish.objects.select_for_update().filter(id__in=qty_by_fish.keys())}
            for fish_id, qty in qty_by_fish.items():
                fish = fishes.get(fish_id)
                if not fish or qty <= 0:
                    continue
                current_stock = int(fish.stock_quantity or 0)
                new_stock = current_stock - qty
                if new_stock < 0:
                    logger.warning(
                        'Order %s attempted to reduce fish %s stock below zero (%s -> %s)',
                        getattr(order, 'order_number', None),
                        fish_id,
                        current_stock,
                        new_stock,
                    )
                    new_stock = 0
                if new_stock != current_stock:
                    fish.stock_quantity = new_stock
                    fish.save(update_fields=['stock_quantity'])
                    deducted = True

    if accessory_items:
        qty_by_accessory = defaultdict(int)
        for item in accessory_items:
            try:
                qty_by_accessory[item.accessory_id] += int(item.quantity or 0)
            except Exception:
                logger.exception('Invalid quantity for accessory item %s on order %s', getattr(item, 'id', None), getattr(order, 'order_number', None))
        if qty_by_accessory:
            accessories = {acc.id: acc for acc in Accessory.objects.select_for_update().filter(id__in=qty_by_accessory.keys())}
            for accessory_id, qty in qty_by_accessory.items():
                accessory = accessories.get(accessory_id)
                if not accessory or qty <= 0:
                    continue
                current_stock = int(accessory.stock_quantity or 0)
                new_stock = current_stock - qty
                if new_stock < 0:
                    logger.warning(
                        'Order %s attempted to reduce accessory %s stock below zero (%s -> %s)',
                        getattr(order, 'order_number', None),
                        accessory_id,
                        current_stock,
                        new_stock,
                    )
                    new_stock = 0
                if new_stock != current_stock:
                    accessory.stock_quantity = new_stock
                    accessory.save(update_fields=['stock_quantity'])
                    deducted = True

    if plant_items:
        qty_by_plant = defaultdict(int)
        for item in plant_items:
            try:
                qty_by_plant[item.plant_id] += int(item.quantity or 0)
            except Exception:
                logger.exception('Invalid quantity for plant item %s on order %s', getattr(item, 'id', None), getattr(order, 'order_number', None))
        if qty_by_plant:
            plants = {pl.id: pl for pl in Plant.objects.select_for_update().filter(id__in=qty_by_plant.keys())}
            for plant_id, qty in qty_by_plant.items():
                plant = plants.get(plant_id)
                if not plant or qty <= 0:
                    continue
                current_stock = int(plant.stock_quantity or 0)
                new_stock = current_stock - qty
                if new_stock < 0:
                    logger.warning(
                        'Order %s attempted to reduce plant %s stock below zero (%s -> %s)',
                        getattr(order, 'order_number', None),
                        plant_id,
                        current_stock,
                        new_stock,
                    )
                    new_stock = 0
                if new_stock != current_stock:
                    plant.stock_quantity = new_stock
                    plant.save(update_fields=['stock_quantity'])
                    deducted = True

    order._inventory_deducted = True
    return deducted


def ensure_order_inventory_deducted(order):
    """Public helper ensuring inventory deduction happens once per order."""
    if transaction.get_connection().in_atomic_block:
        return _ensure_order_inventory_deducted_locked(order)

    with transaction.atomic():
        locked = Order.objects.select_for_update().get(pk=order.pk)
        return _ensure_order_inventory_deducted_locked(locked)


def finalize_order_payment(order, payment_id=None, request=None):
    """Mark an order as paid, deduct stock, clear carts, and trigger invoice send."""
    logger = logging.getLogger(__name__)
    processed = False

    with transaction.atomic():
        locked_order = Order.objects.select_for_update().select_related('user').get(pk=order.pk)

        dirty_fields = []
        if payment_id and locked_order.transaction_id != payment_id:
            locked_order.transaction_id = payment_id
            dirty_fields.append('transaction_id')

        already_processed = locked_order.payment_status == 'paid' and locked_order.status != 'pending'

        if not already_processed:
            try:
                ensure_order_inventory_deducted(locked_order)
            except Exception:
                logger.exception('Inventory deduction failed for order %s', getattr(locked_order, 'order_number', None))
                raise

            if locked_order.payment_status != 'paid':
                locked_order.payment_status = 'paid'
                dirty_fields.append('payment_status')

            if locked_order.status == 'pending':
                locked_order.status = 'processing'
                dirty_fields.append('status')

            processed = True

        if dirty_fields:
            locked_order.save(update_fields=list(dict.fromkeys(dirty_fields)))
        elif processed:
            locked_order.save()

        order = locked_order

    if processed:
        try:
            Cart.objects.filter(user=order.user).delete()
            AccessoryCart.objects.filter(user=order.user).delete()
            PlantCart.objects.filter(user=order.user).delete()
        except Exception:
            logger.exception('Failed to clear cart for order %s', getattr(order, 'order_number', None))

        if request is not None:
            try:
                request.session.pop('applied_coupon_code', None)
            except Exception:
                logger.debug('No coupon session to clear for order %s', getattr(order, 'order_number', None))

        recipient = ''
        if getattr(order, 'user', None):
            recipient = (getattr(order.user, 'email', '') or '').strip()

        if recipient:
            subject = f'Invoice - {settings.SITE_NAME} - {order.order_number}'
            site_base = request.build_absolute_uri('/') if request is not None else getattr(settings, 'SITE_URL', None)
            dispatched = False

            if getattr(settings, 'ORDER_EMAILS_ASYNC', False):
                try:
                    from store.tasks import send_order_email
                    send_order_email.delay(order.id, 'invoice', subject, recipient, site_base)
                    dispatched = True
                except Exception:
                    logger.info('Async invoice queueing failed; falling back to synchronous send for %s', getattr(order, 'order_number', None))

            if not dispatched:
                try:
                    _send_order_email(order, 'invoice', subject, recipient, request=request)
                    dispatched = True
                except Exception:
                    logger.exception('Synchronous invoice send failed for order %s', getattr(order, 'order_number', None))
        else:
            logger.warning('Skipping invoice email - no recipient for order %s', getattr(order, 'order_number', None))

    return processed, order


def generate_invoice_pdf(order):
    """Generate PDF invoice bytes using fpdf2. Returns bytes or None on error."""
    try:
        # Import fpdf locally so missing dependency doesn't break Django startup.
        try:
            from fpdf import FPDF
        except Exception:
            logging.getLogger(__name__).warning('fpdf2 package not available; invoice PDF generation disabled.')
            return None

        # Create a cleaner, more professional invoice layout using fpdf2
        pdf = FPDF(unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Attempt to embed a Unicode-capable TTF (DejaVu Sans) from static/fonts/
        use_unicode_font = False
        try:
            font_path = os.path.join(getattr(settings, 'BASE_DIR', ''), 'static', 'fonts', 'DejaVuSans.ttf')
            if os.path.exists(font_path):
                try:
                    pdf.add_font('DejaVu', '', font_path, uni=True)
                    # Also try bold variant if present (DejaVuSans-Bold.ttf)
                    bold_path = os.path.join(getattr(settings, 'BASE_DIR', ''), 'static', 'fonts', 'DejaVuSans-Bold.ttf')
                    if os.path.exists(bold_path):
                        pdf.add_font('DejaVu', 'B', bold_path, uni=True)
                    use_unicode_font = True
                except Exception:
                    logging.getLogger(__name__).warning('Failed to register DejaVu fonts; falling back to core fonts')
        except Exception:
            use_unicode_font = False

        # Header: logo + company name (use site logo)
        logo_path = os.path.join(getattr(settings, 'BASE_DIR', ''), 'static', 'images', 'logo.jpg')
        if os.path.exists(logo_path):
            try:
                pdf.image(logo_path, x=15, y=10, w=30)
            except Exception:
                pass
        # Move the site name down by 40px. Convert pixels to mm assuming 96 DPI
        # (1 inch = 25.4 mm; px_to_mm = 25.4 / 96). 40px ≈ 10.583 mm.
        try:
            px_to_mm = 25.4 / 96.0
            y_offset_mm = 40 * px_to_mm
        except Exception:
            y_offset_mm = 10.583
        pdf.set_xy(50, 14.5 + y_offset_mm)
        # Choose font family depending on availability
        header_font = 'DejaVu' if use_unicode_font else 'Arial'
        regular_font = 'DejaVu' if use_unicode_font else 'Arial'
        pdf.set_font(header_font, 'B' if use_unicode_font else 'B', 18)
        pdf.set_text_color(11, 83, 148)
        pdf.cell(0, 8, settings.SITE_NAME, ln=True)
        pdf.set_font(regular_font, '', 10)
        pdf.set_text_color(80, 80, 80)
        company_lines = [
            getattr(settings, 'COMPANY_ADDRESS_LINE1', ''),
            getattr(settings, 'COMPANY_ADDRESS_LINE2', ''),
            getattr(settings, 'COMPANY_PHONE', '')
        ]
        for line in [l for l in company_lines if l]:
            pdf.set_x(50)
            pdf.cell(0, 5, line, ln=True)

        pdf.ln(6)

        # Invoice title and meta (right side)
        pdf.set_font(header_font, 'B' if use_unicode_font else 'B', 14)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 8, f'INVOICE', ln=True, align='R')
        pdf.set_font(regular_font, '', 10)
        pdf.cell(0, 5, f'Invoice #: {order.order_number}', ln=True, align='R')
        pdf.cell(0, 5, f'Date: {order.created_at.strftime("%Y-%m-%d")}', ln=True, align='R')

        pdf.ln(4)

        # Billing / Shipping details
        pdf.set_font(header_font, 'B' if use_unicode_font else 'B', 11)
        pdf.cell(95, 6, 'Bill To:', border=0)
        pdf.cell(0, 6, 'Ship To:', border=0, ln=True)

        pdf.set_font(regular_font, '', 10)
        name = f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username
        bill_lines = [name, order.user.email or '']
        ship_lines = (str(order.shipping_address) or '').split('\n') if order.shipping_address else ['N/A']

        max_lines = max(len(bill_lines), len(ship_lines))
        for i in range(max_lines):
            left = bill_lines[i] if i < len(bill_lines) else ''
            right = ship_lines[i] if i < len(ship_lines) else ''
            pdf.cell(95, 5, left, border=0)
            pdf.cell(0, 5, right, border=0, ln=True)

        pdf.ln(6)

        # Table header
        pdf.set_fill_color(242, 246, 251)
        pdf.set_text_color(11, 83, 148)
        pdf.set_draw_color(180, 180, 180)
        pdf.set_line_width(0.4)
        pdf.set_font(regular_font, 'B' if use_unicode_font else 'B', 10)

        col_item_w = 95
        col_qty_w = 20
        col_unit_w = 30
        col_total_w = 30

        th = 8
        pdf.cell(col_item_w, th, 'Item', border=1, fill=True)
        pdf.cell(col_qty_w, th, 'Qty', border=1, fill=True, align='R')
        pdf.cell(col_unit_w, th, 'Unit', border=1, fill=True, align='R')
        pdf.cell(col_total_w, th, 'Total', border=1, fill=True, align='R')
        pdf.ln(th)

        # Table rows — use single-line truncated names for clean alignment
        pdf.set_font(regular_font, '', 10)
        pdf.set_text_color(0, 0, 0)

        def fmt(v):
            try:
                if use_unicode_font:
                    return f'₹{float(v):,.2f}'
                return f'Rs{float(v):,.2f}'
            except Exception:
                return str(v)

        row_fill_toggle = False
        for item in order.items.all():
            row_fill = row_fill_toggle
            row_fill_toggle = not row_fill_toggle
            name = item.fish.name or ''
            name_display = name if len(name) <= 60 else name[:57] + '...'
            fill = 240 if row_fill else 255
            if fill != 255:
                pdf.set_fill_color(fill, fill, fill)
            # Item cell
            pdf.cell(col_item_w, 8, name_display, border=1, fill=(fill != 255))
            pdf.cell(col_qty_w, 8, str(item.quantity), border=1, align='R', fill=(fill != 255))
            pdf.cell(col_unit_w, 8, fmt(item.price), border=1, align='R', fill=(fill != 255))
            pdf.cell(col_total_w, 8, fmt(item.get_total()), border=1, align='R', fill=(fill != 255))
            pdf.ln(8)

        # Accessory items (if any)
        for a in (order.accessory_items.all() if hasattr(order, 'accessory_items') else []):
            row_fill = row_fill_toggle
            row_fill_toggle = not row_fill_toggle
            name = a.accessory.name or ''
            name_display = name if len(name) <= 60 else name[:57] + '...'
            fill = 240 if row_fill else 255
            if fill != 255:
                pdf.set_fill_color(fill, fill, fill)
            pdf.cell(col_item_w, 8, name_display, border=1, fill=(fill != 255))
            pdf.cell(col_qty_w, 8, str(a.quantity), border=1, align='R', fill=(fill != 255))
            pdf.cell(col_unit_w, 8, fmt(a.price), border=1, align='R', fill=(fill != 255))
            pdf.cell(col_total_w, 8, fmt(a.get_total()), border=1, align='R', fill=(fill != 255))
            pdf.ln(8)

        for plant_item in (order.plant_items.all() if hasattr(order, 'plant_items') else []):
            row_fill = row_fill_toggle
            row_fill_toggle = not row_fill_toggle
            name = plant_item.plant.name or ''
            name_display = name if len(name) <= 60 else name[:57] + '...'
            fill = 240 if row_fill else 255
            if fill != 255:
                pdf.set_fill_color(fill, fill, fill)
            pdf.cell(col_item_w, 8, name_display, border=1, fill=(fill != 255))
            pdf.cell(col_qty_w, 8, str(plant_item.quantity), border=1, align='R', fill=(fill != 255))
            pdf.cell(col_unit_w, 8, fmt(plant_item.price), border=1, align='R', fill=(fill != 255))
            pdf.cell(col_total_w, 8, fmt(plant_item.get_total()), border=1, align='R', fill=(fill != 255))
            pdf.ln(8)

        # Totals box
        pdf.ln(6)
        right_x = 15 + col_item_w + col_qty_w
        pdf.set_x(right_x)
        pdf.set_font('Arial', '', 10)
        pdf.cell(col_unit_w, 6, 'Subtotal:', border=0)
        pdf.cell(col_total_w, 6, fmt(order.total_amount), border=0, align='R', ln=True)
        if order.discount_amount and float(order.discount_amount) > 0:
            pdf.set_x(right_x)
            pdf.cell(col_unit_w, 6, 'Discount:', border=0)
            pdf.cell(col_total_w, 6, f"-{fmt(order.discount_amount)}", border=0, align='R', ln=True)
        pdf.set_x(right_x)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(col_unit_w, 8, 'Total:', border=0)
        pdf.cell(col_total_w, 8, fmt(order.final_amount), border=0, align='R', ln=True)

        pdf.ln(8)
        pdf.set_font('Arial', '', 9)
        pdf.multi_cell(0, 5, f'If you have any questions about this invoice, contact us at {getattr(settings, "DEFAULT_FROM_EMAIL", "support@aquafishstore.com")}')

        # Footer
        pdf.set_y(-30)
        pdf.set_font('Arial', 'I', 8)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 5, f'Thank you for shopping with {settings.SITE_NAME}', ln=True, align='C')

        result = pdf.output(dest='S')
        # fpdf2 may return bytes or bytearray; ensure bytes
        if isinstance(result, (bytes, bytearray)):
            return bytes(result)
        try:
            return result.encode('latin-1')
        except Exception:
            return result.encode('utf-8', errors='replace')

        # Logo
        logo_path = os.path.join(getattr(settings, 'BASE_DIR', ''), 'static', 'images', 'hero-betta-fish.png')
        if os.path.exists(logo_path):
            try:
                pdf.image(logo_path, x=10, y=8, w=25)
            except Exception:
                pass

        pdf.set_font('Helvetica', 'B', 16)
        pdf.cell(0, 10, settings.SITE_NAME, ln=True, align='R')
        pdf.ln(4)

        pdf.set_font('Helvetica', '', 10)
        pdf.cell(0, 6, f'Invoice: {order.order_number}', ln=True, align='L')
        pdf.cell(0, 6, f'Date: {order.created_at.strftime("%Y-%m-%d %H:%M")}', ln=True, align='L')
        pdf.ln(4)

        # Bill to
        name = f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 6, 'Bill To:', ln=True)
        pdf.set_font('Helvetica', '', 10)
        pdf.multi_cell(0, 6, name)
        if order.user.email:
            pdf.multi_cell(0, 6, order.user.email)
        if order.shipping_address:
            for line in str(order.shipping_address).split('\n'):
                if line.strip():
                    pdf.multi_cell(0, 6, line.strip())

        pdf.ln(6)

        # Table header
        pdf.set_fill_color(242, 246, 251)
        pdf.set_text_color(11, 83, 148)
        pdf.set_draw_color(200,200,200)
        pdf.set_font('Helvetica', 'B', 10)
        th = 8
        pdf.cell(90, th, 'Item', border=1, fill=True)
        pdf.cell(20, th, 'Qty', border=1, fill=True, align='R')
        pdf.cell(30, th, 'Unit', border=1, fill=True, align='R')
        pdf.cell(30, th, 'Total', border=1, fill=True, align='R')
        pdf.ln(th)

        # Table rows
        pdf.set_font('Helvetica', '', 10)
        pdf.set_text_color(0,0,0)
        def fmt(v):
            try:
                return f'₹{float(v):,.2f}'
            except Exception:
                return str(v)

        for item in order.items.all():
            pdf.cell(90, 7, item.fish.name[:60], border=1)
            pdf.cell(20, 7, str(item.quantity), border=1, align='R')
            pdf.cell(30, 7, fmt(item.price), border=1, align='R')
            pdf.cell(30, 7, fmt(item.get_total()), border=1, align='R')
            pdf.ln(7)

        # Totals
        pdf.ln(4)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(140, 6, '', border=0)
        pdf.cell(30, 6, 'Subtotal:', border=0)
        pdf.cell(30, 6, fmt(order.total_amount), border=0, align='R')
        pdf.ln(6)
        if order.discount_amount and float(order.discount_amount) > 0:
            pdf.cell(140, 6, '', border=0)
            pdf.cell(30, 6, 'Discount:', border=0)
            pdf.cell(30, 6, f"-{fmt(order.discount_amount)}", border=0, align='R')
            pdf.ln(6)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(140, 8, '', border=0)
        pdf.cell(30, 8, 'Total:', border=0)
        pdf.cell(30, 8, fmt(order.final_amount), border=0, align='R')
        pdf.ln(12)

        pdf.set_font('Helvetica', '', 9)
        pdf.multi_cell(0, 6, f'Thank you for shopping with {settings.SITE_NAME}')

        out = pdf.output(dest='S').encode('latin-1')
        return out
    except Exception:
        logging.getLogger(__name__).exception('Error generating invoice PDF (fpdf) for order %s', getattr(order, 'order_number', None))
        return None


def resend_otp_view(request):
    pending = request.session.get('pending_registration')
    if not pending:
        messages.error(request, 'No registration in progress. Please register first.')
        return redirect('register')

    if request.method == 'POST':
        otp_code = OTP.generate_otp()
        request.session['pending_registration_otp'] = otp_code
        request.session['pending_registration_time'] = timezone.now().isoformat()

        send_mail(
            f'Email Verification OTP - {settings.SITE_NAME}',
            f"Your OTP for email verification is: {otp_code}\n\nThis OTP will expire in 5 minutes.",
            settings.DEFAULT_FROM_EMAIL or 'noreply@aquafishstore.com',
            [pending.get('email')],
            fail_silently=False,
        )

        messages.success(request, 'A new OTP has been sent to your email.')
        return redirect('verify_otp')

    return render(request, 'store/resend_otp.html', {'pending_email': pending.get('email')})


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user:
            if user.is_blocked:
                messages.error(request, 'Your account has been blocked. Please contact admin.')
                return render(request, 'store/login.html')
            
            if not user.email_verified:
                messages.error(request, 'Please verify your email first.')
                # Registration OTP flow is session-based; just send the user to verification page.
                # Note: If there's no pending registration session, the page will prompt them to register again.
                return redirect('verify_otp')
            
            login(request, user)
            _merge_guest_cart_into_user(request, user)
            
            # Redirect based on role
            if user.role == 'admin':
                return redirect('admin_dashboard')
            elif user.role == 'staff':
                return redirect('staff_dashboard')
            else:
                return redirect('home')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'store/login.html')


@require_POST
def ajax_login_view(request):
    """AJAX login endpoint used by the inline login modal.

    Expects `username` and `password` in POST data and returns JSON.
    """
    username = request.POST.get('username')
    password = request.POST.get('password')
    if not username or not password:
        return JsonResponse({'success': False, 'message': 'Please provide username and password'}, status=400)

    user = authenticate(request, username=username, password=password)
    if user is None:
        return JsonResponse({'success': False, 'message': 'Invalid credentials'}, status=400)
    if getattr(user, 'is_blocked', False):
        return JsonResponse({'success': False, 'message': 'Your account is blocked'}, status=403)

    login(request, user)
    _merge_guest_cart_into_user(request, user)
    return JsonResponse({'success': True, 'message': 'Logged in'})


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


def test_email_view(request):
    """DEBUG-only endpoint to test email delivery.
    Usage: /test-email/?to=address@example.com
    Sends a simple email and returns backend info or error details.
    """
    if not settings.DEBUG:
        return HttpResponse(status=404)
    to_addr = request.GET.get('to') or (request.user.email if request.user.is_authenticated else None)
    if not to_addr:
        return HttpResponse("Provide ?to=email@example.com or login with an email.", status=400)
    try:
        send_mail(
            f'{settings.SITE_NAME} - Test Email',
            f'This is a test email from {settings.SITE_NAME}. If you received this, SMTP is configured correctly.',
            settings.DEFAULT_FROM_EMAIL or 'noreply@aquafishstore.local',
            [to_addr],
            fail_silently=False,
        )
        return HttpResponse(f"Sent test email to {to_addr} using backend {settings.EMAIL_BACKEND}.")
    except Exception as e:
        return HttpResponse(f"Email send error: {e}", status=500)


def email_config_view(request):
    """DEBUG diagnostic: shows current email settings and attempts SMTP connectivity."""
    if not settings.DEBUG:
        return HttpResponse(status=404)
    info = {
        'EMAIL_BACKEND': settings.EMAIL_BACKEND,
        'EMAIL_HOST': getattr(settings, 'EMAIL_HOST', None),
        'EMAIL_PORT': getattr(settings, 'EMAIL_PORT', None),
        'EMAIL_USE_TLS': getattr(settings, 'EMAIL_USE_TLS', None),
        'EMAIL_USE_SSL': getattr(settings, 'EMAIL_USE_SSL', None),
        'EMAIL_HOST_USER': getattr(settings, 'EMAIL_HOST_USER', None),
        'DEFAULT_FROM_EMAIL': settings.DEFAULT_FROM_EMAIL,
        'EMAIL_TIMEOUT': getattr(settings, 'EMAIL_TIMEOUT', None),
    }
    connectivity = 'skip (console backend)'
    error = None
    if info['EMAIL_BACKEND'] != 'django.core.mail.backends.console.EmailBackend' and info['EMAIL_HOST'] and info['EMAIL_PORT']:
        try:
            if info['EMAIL_USE_SSL']:
                server = smtplib.SMTP_SSL(host=info['EMAIL_HOST'], port=info['EMAIL_PORT'], timeout=info['EMAIL_TIMEOUT'])
            else:
                server = smtplib.SMTP(host=info['EMAIL_HOST'], port=info['EMAIL_PORT'], timeout=info['EMAIL_TIMEOUT'])
                if info['EMAIL_USE_TLS']:
                    server.starttls()
            server.quit()
            connectivity = 'ok'
        except (socket.error, smtplib.SMTPException) as e:
            connectivity = 'fail'
            error = str(e)

    html = ["<h2>Email Configuration Diagnostic</h2>"]
    html.append("<table border='1' cellpadding='5' style='border-collapse:collapse;font-family:monospace;'>")
    for k, v in info.items():
        html.append(f"<tr><th>{k}</th><td>{v}</td></tr>")
    html.append(f"<tr><th>SMTP Connectivity</th><td>{connectivity}</td></tr>")
    if error:
        html.append(f"<tr><th>Error</th><td style='color:red'>{error}</td></tr>")
    html.append("</table>")
    html.append("<p>Use /test-email/?to=address to send a test message.</p>")
    return HttpResponse(''.join(html))


def forgot_password_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        user = CustomUser.objects.filter(email=email).first()
        
        if user:
            if user.role == 'admin':
                messages.error(request, 'Admin users cannot use forgot password. Please contact system administrator.')
                return render(request, 'store/forgot_password.html')
            
            otp_code = OTP.generate_otp()
            OTP.objects.create(user=user, otp_code=otp_code)
            
            send_mail(
                f'Password Reset OTP - {settings.SITE_NAME}',
                f'Your OTP for password reset is: {otp_code}\n\nThis OTP will expire in 5 minutes.',
                settings.DEFAULT_FROM_EMAIL or 'noreply@aquafishstore.com',
                [user.email],
                fail_silently=False,
            )
            
            messages.success(request, 'OTP has been sent to your email.')
            return redirect('reset_password', user_id=user.id)
        else:
            messages.error(request, 'No account found with this email.')
    
    return render(request, 'store/forgot_password.html')


def reset_password_view(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == 'POST':
        otp_code = request.POST.get('otp_code')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        otp_obj = OTP.objects.filter(user=user, is_used=False).order_by('-created_at').first()
        
        if otp_obj and otp_obj.otp_code == otp_code:
            if otp_obj.is_expired():
                messages.error(request, 'OTP has expired. Please request a new one.')
                return redirect('forgot_password')
            
            if new_password == confirm_password:
                otp_obj.is_used = True
                otp_obj.save()
                user.set_password(new_password)
                user.save()
                messages.success(request, 'Password reset successful! Please login with your new password.')
                return redirect('login')
            else:
                messages.error(request, 'Passwords do not match.')
        else:
            messages.error(request, 'Invalid OTP.')
    
    return render(request, 'store/reset_password.html', {'user': user})


# Home/Customer Views
def home_view(request):
    categories = Category.objects.all()[:6]
    fishes = Fish.objects.filter(is_available=True, stock_quantity__gt=0, is_featured=True)[:8]
    # Limited offers currently active and marked to show on homepage
    now = timezone.now()
    limited_offers_qs = LimitedOffer.objects.filter(
        is_active=True,
        show_on_homepage=True,
    ).filter(
        Q(start_time__isnull=True, end_time__isnull=True) |  # No scheduling set
        Q(start_time__isnull=True, end_time__gte=now) |       # Only end_time set
        Q(start_time__lte=now, end_time__isnull=True) |       # Only start_time set
        Q(start_time__lte=now, end_time__gte=now)             # Both set and within range
    ).order_by('end_time')[:4]

    # Convert LimitedOffer model instances into plain dicts so template can treat
    # both model-backed offers and combo-generated offers uniformly.
    limited_offers = []
    for offer in limited_offers_qs:
        entry = {
            'title': offer.title,
            'description': offer.description,
            'discount_text': offer.discount_text,
            'image': (offer.image.url if getattr(offer, 'image', None) else None),
            'bg_color': offer.bg_color,
            'fish': offer.fish,
            'start_time': offer.start_time,
            'end_time': offer.end_time,
            'is_combo': False,
            'combo_id': None,
            'images': [],
        }

        if getattr(offer, 'combo', None):
            combo = offer.combo
            items = list(combo.items.select_related('fish').all())
            images = []
            for it in items:
                try:
                    if it.fish.image:
                        images.append(it.fish.image.url)
                except Exception:
                    continue
                if len(images) >= 3:
                    break

            entry['is_combo'] = True
            entry['combo_id'] = combo.id
            entry['images'] = images
            # Prefer combo-derived discount text if available
            if combo.bundle_price:
                original_total = sum((it.fish.price or 0) * it.quantity for it in items)
                try:
                    savings = float(original_total) - float(combo.bundle_price)
                except Exception:
                    savings = 0
                if savings > 0:
                    entry['discount_text'] = f"Save ₹{int(savings)}"

        limited_offers.append(entry)

    # Build a list of all active combos (used elsewhere) and a separate
    # `combo_deals` list containing only combos marked for homepage display.
    combos = ComboOffer.objects.filter(is_active=True)
    combo_offers = []
    for combo in combos:
        # Compute approximate savings if bundle_price set
        items = list(combo.items.select_related('fish').all())
        if not items:
            continue
        original_total = sum((it.fish.price or 0) * it.quantity for it in items)
        bundle_price = combo.bundle_price
        if bundle_price:
            try:
                savings = float(original_total) - float(bundle_price)
            except Exception:
                savings = 0
        else:
            savings = 0

        if savings > 0:
            # Prefer a rupee saving label
            discount_text = f"Save ₹{int(savings)}"
        else:
            discount_text = 'Bundle Deal'

        first_fish = items[0].fish if items else None
        # Collect up to three image URLs from the combo fishes for a thumbnail strip
        images = []
        for it in items:
            try:
                if it.fish.image:
                    images.append(it.fish.image.url)
            except Exception:
                continue
            if len(images) >= 3:
                break

        combo_offers.append({
            'title': combo.title,
            'description': combo.description,
            'discount_text': discount_text,
            'image': None,
            'bg_color': None,
            'fish': first_fish,
            'start_time': now,
            'end_time': now + timedelta(days=7),
            'is_combo': True,
            'combo_id': combo.id,
            'images': images,
            'item_names': [it.fish.name for it in items if getattr(it, 'fish', None)],
        })
    # Build `combo_deals` for homepage (only combos explicitly marked and not banners)
    combo_deals_qs = ComboOffer.objects.filter(is_active=True, show_on_homepage=True, show_as_banner=False).prefetch_related('items__fish').order_by('-created_at')
    combo_deals = []
    # Map to the same dict shape used for limited_offers so templates can use
    # a consistent structure when rendering banners/cards.
    combo_map = {c['combo_id']: c for c in combo_offers}
    for combo in combo_deals_qs:
        # Try to reuse the already-computed entry if present
        existing = next((c for c in combo_offers if c['combo_id'] == combo.id), None)
        if existing:
            combo_deals.append(existing)
        else:
            # fallback mapping if not in combo_offers
            items = list(combo.items.select_related('fish').all())
            original_total = sum((it.fish.price or 0) * it.quantity for it in items) if items else 0
            bundle_price = combo.bundle_price
            try:
                savings = float(original_total) - float(bundle_price) if bundle_price else 0
            except Exception:
                savings = 0
            discount_text = f"Save ₹{int(savings)}" if savings > 0 else 'Bundle Deal'
            images = []
            for it in items:
                try:
                    if it.fish.image:
                        images.append(it.fish.image.url)
                except Exception:
                    continue
                if len(images) >= 3:
                    break
            combo_deals.append({
                'title': combo.title,
                'description': combo.description,
                'discount_text': discount_text,
                'image': None,
                'bg_color': None,
                'fish': items[0].fish if items else None,
                'start_time': now,
                'end_time': now + timedelta(days=7),
                'is_combo': True,
                'combo_id': combo.id,
                'images': images,
                'item_names': [it.fish.name for it in items if getattr(it, 'fish', None)],
            })
    # Build `combo_banners` for homepage (combos marked to show as banner)
    combo_banners = []
    combo_banners_qs = ComboOffer.objects.filter(is_active=True, show_as_banner=True).prefetch_related('items__fish').order_by('-created_at')
    for combo in combo_banners_qs:
        img = None
        try:
            if getattr(combo, 'banner_image', None):
                img = combo.banner_image.url
        except Exception:
            img = None
        combo_banners.append({
            'title': combo.title,
            'description': combo.description,
            'image': img,
            'combo_id': combo.id,
            'item_names': [it.fish.name for it in combo.items.all() if getattr(it, 'fish', None)],
        })
    # Provide a hero_items list for the homepage: prefer explicit combo banners,
    # but fall back to combo_deals (which may have images in 'images' or 'fish')
    hero_items = []
    if combo_banners:
        hero_items = combo_banners
    else:
        for cd in combo_deals:
            img = cd.get('image')
            if not img:
                imgs = cd.get('images') or []
                if imgs:
                    img = imgs[0]
                else:
                    fish = cd.get('fish')
                    try:
                        img = fish.image.url if fish and getattr(fish, 'image', None) else None
                    except Exception:
                        img = None
            hero_items.append({
                'title': cd.get('title'),
                'description': cd.get('description'),
                'image': img,
                'combo_id': cd.get('combo_id'),
                'discount_text': cd.get('discount_text'),
                'item_names': cd.get('item_names', []),
            })
    first_category = Category.objects.first() if Category.objects.exists() else None
    reviews = (Review.objects.filter(approved=True)
               .select_related('order', 'user')
               .prefetch_related('order__items', 'order__items__fish')[:10])
    return render(request, 'store/home.html', {
        'categories': categories,
        'fishes': fishes,
        'first_category': first_category,
        'reviews': reviews,
        'limited_offers': limited_offers,
        'combo_offers': combo_offers,
        'combo_deals': combo_deals,
        'combo_banners': combo_banners,
        'hero_items': hero_items,
    })

@login_required
@user_passes_test(is_admin)
def admin_limited_offers_view(request):
    # Avoid ordering by `created_at` in case DB schema hasn't been migrated yet.
    offers = LimitedOffer.objects.all().order_by('-id')
    return render(request, 'store/admin/limited_offers.html', {'offers': offers})

@login_required
@user_passes_test(is_admin)
def admin_add_limited_offer_view(request):
    if request.method == 'POST':
        form = LimitedOfferForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Limited offer created successfully.')
            return redirect('admin_limited_offers')
    else:
        form = LimitedOfferForm()
    return render(request, 'store/admin/add_limited_offer.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def admin_edit_limited_offer_view(request, offer_id):
    offer = get_object_or_404(LimitedOffer, id=offer_id)
    if request.method == 'POST':
        form = LimitedOfferForm(request.POST, request.FILES, instance=offer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Limited offer updated successfully.')
            return redirect('admin_limited_offers')
    else:
        form = LimitedOfferForm(instance=offer)
    return render(request, 'store/admin/add_limited_offer.html', {'form': form, 'offer': offer})

@login_required
@user_passes_test(is_admin)
def admin_toggle_limited_offer_view(request, offer_id):
    offer = get_object_or_404(LimitedOffer, id=offer_id)
    offer.is_active = not offer.is_active
    offer.save(update_fields=['is_active'])
    messages.success(request, f"Offer {'activated' if offer.is_active else 'deactivated'} successfully.")
    return redirect('admin_limited_offers')


@login_required
@user_passes_test(is_admin)
def admin_toggle_combo_banners_view(request):
    """Toggle showing combo deals on the homepage for all active combos.

    If any active combo is currently set to show_on_homepage, hide all; otherwise show all.
    """
    from .models import ComboOffer
    combos = ComboOffer.objects.filter(is_active=True)
    if combos.filter(show_on_homepage=True).exists():
        combos.update(show_on_homepage=False)
        messages.success(request, 'Combo banners hidden on the homepage.')
    else:
        combos.update(show_on_homepage=True)
        messages.success(request, 'Combo banners will be shown on the homepage.')
    return redirect('profile')


@login_required
@user_passes_test(is_admin)
def admin_combo_deals_view(request):
    """Show a page where admin can select which combos should appear on the homepage.

    GET: render list of combos with checkboxes.
    POST: expect `combo_ids` (list of ids) — set show_on_homepage True for listed ids, False for others.
    """
    from .models import ComboOffer
    from .forms import ComboDealsForm

    if request.method == 'POST':
        # Accept POST with possible file uploads for banners
        form = ComboDealsForm(request.POST, request.FILES)
        if form.is_valid():
            selected = form.cleaned_data.get('combos') or ComboOffer.objects.none()
            # First, set show_on_homepage False for all active combos
            active_combos = ComboOffer.objects.filter(is_active=True)
            active_combos.update(show_on_homepage=False)

            # Apply show_on_homepage to selected combos unless they are set to show as banner
            selected_ids = [c.id for c in selected] if selected.exists() else []
            if selected_ids:
                ComboOffer.objects.filter(id__in=selected_ids, is_active=True).update(show_on_homepage=True)

            # Now process per-combo banner flags and uploaded files
            for combo in ComboOffer.objects.filter(is_active=True):
                sid = str(combo.id)
                # Banner toggle input name: show_as_banner_<id>
                show_banner_val = request.POST.get(f'show_as_banner_{sid}')
                requested_show_as_banner = True if show_banner_val in ('1', 'on', 'true', 'True') else False

                # Handle uploaded file for banner image: input name banner_<id>
                uploaded = request.FILES.get(f'banner_{sid}')
                # Handle clear image checkbox: name clear_banner_<id>
                clear_flag = request.POST.get(f'clear_banner_{sid}')

                # If an uploaded file is provided, prefer it. Otherwise if clear flag is set, delete existing image.
                if uploaded:
                    combo.banner_image = uploaded
                else:
                    if clear_flag:
                        try:
                            if combo.banner_image:
                                combo.banner_image.delete(save=False)
                        except Exception:
                            pass
                        combo.banner_image = None

                # Only enable banner mode if an image exists (either previously uploaded or in this POST)
                has_image = bool(getattr(combo, 'banner_image', None))
                show_as_banner = requested_show_as_banner and has_image

                if show_as_banner:
                    # Ensure it will not also show as a card
                    combo.show_on_homepage = False
                else:
                    # If not banner, preserve show_on_homepage from selected list
                    combo.show_on_homepage = combo.id in selected_ids

                # If the admin requested banner but no image was provided, warn them
                if requested_show_as_banner and not has_image:
                    messages.warning(request, f"Combo '{combo.title}' cannot be enabled as a banner without an image.")

                combo.show_as_banner = show_as_banner
                combo.save()

            messages.success(request, 'Combo homepage visibility and banners updated.')
            return redirect('admin_combo_deals')
        # If form is invalid, fall through to re-render with errors
        combos = ComboOffer.objects.order_by('-created_at').all()
        cards_enabled = ComboOffer.objects.filter(is_active=True, show_on_homepage=True).exists()
        banners_enabled = ComboOffer.objects.filter(is_active=True, show_as_banner=True).exists()
        return render(
            request,
            'store/admin/combo_deals.html',
            {
                'combos': combos,
                'form': form,
                'cards_enabled': cards_enabled,
                'banners_enabled': banners_enabled,
            },
        )
    else:
        # Initialize form with currently selected combos
        initial_qs = ComboOffer.objects.filter(is_active=True, show_on_homepage=True)
        form = ComboDealsForm(initial={'combos': initial_qs})

        combos = ComboOffer.objects.order_by('-created_at').all()
        cards_enabled = ComboOffer.objects.filter(is_active=True, show_on_homepage=True).exists()
        banners_enabled = ComboOffer.objects.filter(is_active=True, show_as_banner=True).exists()
        return render(request, 'store/admin/combo_deals.html', {'combos': combos, 'form': form, 'cards_enabled': cards_enabled, 'banners_enabled': banners_enabled})

    # Safety fallback: ensure an HttpResponse is always returned even if
    # unforeseen code paths are hit during reloads or concurrent edits.
    try:
        fallback_combos = ComboOffer.objects.order_by('-created_at').all()
        fallback_form = ComboDealsForm()
        fallback_cards = ComboOffer.objects.filter(is_active=True, show_on_homepage=True).exists()
        fallback_banners = ComboOffer.objects.filter(is_active=True, show_as_banner=True).exists()
        return render(request, 'store/admin/combo_deals.html', {
            'combos': fallback_combos,
            'form': fallback_form,
            'cards_enabled': fallback_cards,
            'banners_enabled': fallback_banners,
        })
    except Exception:
        # As an absolute last resort, return a simple HttpResponse so Django
        # doesn't raise a ValueError about a missing response. Keep it minimal
        # to avoid exposing internals.
        from django.http import HttpResponse
        return HttpResponse('Combo deals page temporarily unavailable', status=503)


@login_required
@user_passes_test(is_admin)
def ajax_toggle_combo_banner(request):
    """AJAX endpoint to toggle show_as_banner for a combo.

    Expects JSON body: { combo_id: <id>, value: 0|1 }
    """
    import json
    from django.http import JsonResponse
    from .models import ComboOffer

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)

    # parse JSON body, fall back to POST form data
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        payload = {}

    combo_id = payload.get('combo_id') or request.POST.get('combo_id')
    val = payload.get('value') or request.POST.get('value')

    if not combo_id:
        return JsonResponse({'success': False, 'message': 'combo_id required'}, status=400)

    try:
        combo = ComboOffer.objects.get(id=int(combo_id))
    except Exception:
        return JsonResponse({'success': False, 'message': 'combo not found'}, status=404)

    show_as_banner = True if str(val) in ('1', 'true', 'True', 'on') else False
    combo.show_as_banner = show_as_banner
    if show_as_banner:
        combo.show_on_homepage = False
    combo.save()

    return JsonResponse({'success': True, 'show_as_banner': combo.show_as_banner})


@login_required
@user_passes_test(is_admin)
def ajax_toggle_combo_banners(request):
    """AJAX endpoint to toggle showing combo banners globally.

    Enables `show_as_banner` for combos that have a `banner_image` and disables `show_on_homepage`.
    When disabling, clears `show_as_banner` for all combos.
    """
    from django.http import JsonResponse
    from .models import ComboOffer

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)

    active = ComboOffer.objects.filter(is_active=True)
    enabled = active.filter(show_as_banner=True).exists()
    if enabled:
        # disable banners
        active.update(show_as_banner=False)
        new_state = False
    else:
        # enable banners for combos which have a banner_image and ensure cards are hidden
        active.update(show_on_homepage=False)
        active_with_img = active.exclude(banner_image__isnull=True).exclude(banner_image__exact='')
        active_with_img.update(show_as_banner=True)
        new_state = True

    return JsonResponse({'success': True, 'enabled': new_state})

@login_required
@user_passes_test(is_admin)
def admin_delete_limited_offer_view(request, offer_id):
    offer = get_object_or_404(LimitedOffer, id=offer_id)
    offer.delete()
    messages.success(request, 'Offer deleted successfully.')
    return redirect('admin_limited_offers')


@login_required
@user_passes_test(is_admin)
def ajax_toggle_combo_cards(request):
    """AJAX endpoint to toggle showing combo cards globally.

    Toggles `show_on_homepage` for all active combos and clears any `show_as_banner` flags.
    """
    from django.http import JsonResponse
    from .models import ComboOffer

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)

    active = ComboOffer.objects.filter(is_active=True)
    enabled = active.filter(show_on_homepage=True).exists()
    if enabled:
        # disable cards
        active.update(show_on_homepage=False)
        new_state = False
    else:
        # enable cards and clear banner flags
        active.update(show_on_homepage=True, show_as_banner=False)
        new_state = True

    return JsonResponse({'success': True, 'enabled': new_state})


def about_view(request):
    services = Service.objects.filter(is_active=True)
    contact = ContactInfo.objects.first()
    has_gallery = False
    map_url = None
    if contact:
        try:
            has_gallery = contact.gallery_media.exists()
        except Exception:
            has_gallery = False
        # Prefer a proper Google Maps embed URL if provided
        raw = (contact.map_embed_url or '').strip()
        if raw and 'google.com/maps/embed' in raw:
            map_url = raw
        else:
            # Fallback: build a query-based embed from the address
            parts = [
                contact.address_line1 or '',
                contact.address_line2 or '',
                contact.city or '',
                contact.state or '',
                contact.postal_code or '',
                contact.country or '',
            ]
            full_addr = ' '.join([p for p in parts if p]).strip()
            if full_addr:
                map_url = f"https://www.google.com/maps?q={quote_plus(full_addr)}&output=embed"
    return render(request, 'store/about.html', {
        'services': services,
        'contact': contact,
        'has_gallery': has_gallery,
        'map_url': map_url,
    })


def terms_and_conditions_view(request):
    return render(request, 'store/terms_and_conditions.html')


def privacy_policy_view(request):
    return render(request, 'store/privacy_policy.html')


def return_policy_view(request):
    return render(request, 'store/return_policy.html')


def cookie_policy_view(request):
    return render(request, 'store/cookie_policy.html')


def terms_of_use_view(request):
    return render(request, 'store/terms_of_use.html')


def help_center_view(request):
    return render(request, 'store/help_center.html')


def blog_list_view(request):
    """Public blog list showing published posts."""
    from .models import BlogPost
    posts = BlogPost.objects.filter(is_published=True).order_by('-published_at')
    paginator = Paginator(posts, 10)
    page = request.GET.get('page', 1)
    try:
        posts_page = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        posts_page = paginator.page(1)
    return render(request, 'store/blog_list.html', {'posts': posts_page})


def blog_detail_view(request, slug):
    from .models import BlogPost
    post = get_object_or_404(BlogPost, slug=slug, is_published=True)
    # Provide recent published posts to the template (avoid ORM calls from templates)
    recent_posts = BlogPost.objects.filter(is_published=True).order_by('-published_at')[:5]
    return render(request, 'store/blog_detail.html', {'post': post, 'recent_posts': recent_posts})


def customer_fish_list_view(request):
    fishes = Fish.objects.filter(is_available=True, stock_quantity__gt=0)
    categories = Category.objects.filter(category_type='fish').order_by('name')
    breeds = Breed.objects.all()
    
    # Search and filter
    search_query = request.GET.get('search', '')
    category_filter = request.GET.get('category', '')
    breed_filter = request.GET.get('breed', '')
    
    if search_query:
        fishes = fishes.filter(
            Q(name__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(breed__name__icontains=search_query)
        )
    
    if category_filter:
        fishes = fishes.filter(category_id=category_filter)
    
    if breed_filter:
        fishes = fishes.filter(breed_id=breed_filter)
    
    # If AJAX request, return only the rendered grid partial for live search
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render(request, 'store/customer/_fish_grid.html', {'fishes': fishes}).content.decode('utf-8')
        return JsonResponse({'html': html})

    return render(request, 'store/customer/fish_list.html', {
        'fishes': fishes,
        'categories': categories,
        'breeds': breeds,
        'search_query': search_query,
        'category_filter': category_filter,
        'breed_filter': breed_filter,
    })


def customer_plants_view(request):
    plants_qs = Plant.objects.filter(is_active=True, stock_quantity__gt=0).select_related('category')
    categories = Category.objects.filter(category_type='plant').order_by('name')

    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '').strip()

    if search_query:
        plants_qs = plants_qs.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if category_filter and category_filter.isdigit():
        plants_qs = plants_qs.filter(category_id=int(category_filter))

    paginator = Paginator(plants_qs, 12)
    page_number = request.GET.get('page')
    try:
        plants = paginator.page(page_number or 1)
    except PageNotAnInteger:
        plants = paginator.page(1)
    except EmptyPage:
        plants = paginator.page(paginator.num_pages)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render_to_string(
            'store/customer/partials/plant_grid.html',
            {
                'plants': plants,
                'search_query': search_query,
                'category_filter': category_filter,
            },
            request=request,
        )
        return JsonResponse({'html': html})

    return render(request, 'store/customer/plants.html', {
        'plants': plants,
        'categories': categories,
        'search_query': search_query,
        'category_filter': category_filter,
    })


def fish_detail_view(request, fish_id):
    fish = get_object_or_404(Fish, id=fish_id, is_available=True, stock_quantity__gt=0)
    # Collect up to 5 images and 2 videos
    images = list(FishMedia.objects.filter(fish=fish, media_type='image')[:5])
    videos = list(FishMedia.objects.filter(fish=fish, media_type='video')[:2])
    media = images + videos
    return render(request, 'store/customer/fish_detail.html', {'fish': fish, 'media': media})


def plant_detail_view(request, plant_id):
    plant = get_object_or_404(Plant, id=plant_id, is_active=True, stock_quantity__gt=0)
    gallery_items = list(PlantMedia.objects.filter(plant=plant))

    primary_image_url = None
    if getattr(plant, 'image', None):
        try:
            primary_image_url = plant.image.url
        except Exception:
            primary_image_url = None
    if not primary_image_url and gallery_items:
        primary_image_url = gallery_items[0].source

    modal_images = []
    if primary_image_url:
        modal_images.append({'url': primary_image_url, 'title': plant.name})
    # Include remaining gallery images, skipping duplicates of the primary image URL
    for item in gallery_items:
        src = item.source
        if not src:
            continue
        if primary_image_url and src == primary_image_url:
            continue
        modal_images.append({'url': src, 'title': item.title or plant.name})

    return render(request, 'store/customer/plant_detail.html', {
        'plant': plant,
        'primary_image_url': primary_image_url,
        'modal_images': modal_images,
        'gallery_items': gallery_items,
    })


def customer_accessories_view(request):
    """List available accessories for customers"""
    accessories_qs = Accessory.objects.filter(is_active=True, stock_quantity__gt=0)
    # query params
    q = request.GET.get('q', '').strip()
    category_id = request.GET.get('category', '').strip()
    categories = Category.objects.filter(category_type='accessory').order_by('name')

    if q:
        accessories_qs = accessories_qs.filter(
            Q(name__icontains=q) | Q(description__icontains=q) | Q(category__name__icontains=q)
        )

    if category_id:
        try:
            accessories_qs = accessories_qs.filter(category_id=int(category_id))
        except ValueError:
            pass

    # Default ordering: newest first if available
    if hasattr(Accessory, 'created_at'):
        accessories_qs = accessories_qs.order_by('-created_at')
    else:
        accessories_qs = accessories_qs.order_by('-id')

    # No pagination — return the full accessory list as requested
    # If AJAX request, return partial rendering of the full grid
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render(request, 'store/customer/_accessory_grid.html', {'accessories': accessories_qs}).content.decode('utf-8')
        return JsonResponse({'html': html})

    return render(request, 'store/customer/accessories.html', {
        'accessories': accessories_qs,
        'search_query': q,
        'categories': categories,
        'selected_category': category_id,
    })


def accessory_detail_view(request, accessory_id):
    accessory = get_object_or_404(Accessory, id=accessory_id, is_active=True, stock_quantity__gt=0)
    return render(request, 'store/customer/accessory_detail.html', {'accessory': accessory})


# Admin: Manage Fish Media
@login_required
@user_passes_test(is_admin)
def admin_fish_media_view(request, fish_id):
    fish = get_object_or_404(Fish, id=fish_id)
    media_items = FishMedia.objects.filter(fish=fish)

    if request.method == 'POST':
        form = FishMediaForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.fish = fish
            # Basic validation: require either file or external_url
            if not obj.file and not obj.external_url:
                messages.error(request, 'Please provide a file or an external URL.')
            else:
                obj.save()
                messages.success(request, 'Media added successfully.')
                return redirect('admin_fish_media', fish_id=fish.id)
    else:
        form = FishMediaForm()

    return render(request, 'store/admin/fish_media.html', {
        'fish': fish,
        'media_items': media_items,
        'form': form,
    })


@login_required
@user_passes_test(is_admin)
def admin_delete_fish_media_view(request, media_id):
    media = get_object_or_404(FishMedia, id=media_id)
    fish_id = media.fish.id
    media.delete()
    messages.success(request, 'Media deleted successfully.')
    return redirect('admin_fish_media', fish_id=fish_id)


@login_required
@user_passes_test(is_admin)
def admin_edit_fish_media_view(request, media_id):
    """Edit an existing FishMedia item (title, order, file, external URL, type)."""
    media = get_object_or_404(FishMedia, id=media_id)
    fish = media.fish

    if request.method == 'POST':
        form = FishMediaForm(request.POST, request.FILES, instance=media)
        if form.is_valid():
            obj = form.save(commit=False)
            # Require either a file or external URL (for videos). Allow images without external_url.
            if not obj.file and not obj.external_url:
                messages.error(request, 'Please provide a file or an external URL.')
            else:
                obj.save()
                messages.success(request, 'Media updated successfully.')
                return redirect('admin_fish_media', fish_id=fish.id)
    else:
        form = FishMediaForm(instance=media)

    return render(request, 'store/admin/edit_fish_media.html', {
        'fish': fish,
        'media': media,
        'form': form,
    })


@login_required
@user_passes_test(is_admin)
def admin_plant_media_view(request, plant_id):
    plant = get_object_or_404(Plant, id=plant_id)
    media_items = PlantMedia.objects.filter(plant=plant)

    if request.method == 'POST':
        form = PlantMediaForm(request.POST, request.FILES)
        if form.is_valid():
            media = form.save(commit=False)
            media.plant = plant
            media.save()
            messages.success(request, 'Image added to gallery.')
            return redirect('admin_plant_media', plant_id=plant.id)
    else:
        form = PlantMediaForm()

    return render(request, 'store/admin/plant_media.html', {
        'plant': plant,
        'media_items': media_items,
        'form': form,
    })


@login_required
@user_passes_test(is_admin)
def admin_delete_plant_media_view(request, media_id):
    media = get_object_or_404(PlantMedia, id=media_id)
    plant_id = media.plant_id
    media.delete()
    messages.success(request, 'Gallery image deleted.')
    return redirect('admin_plant_media', plant_id=plant_id)


@login_required
@user_passes_test(is_admin)
def admin_edit_plant_media_view(request, media_id):
    media = get_object_or_404(PlantMedia, id=media_id)
    plant = media.plant

    if request.method == 'POST':
        form = PlantMediaForm(request.POST, request.FILES, instance=media)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gallery image updated.')
            return redirect('admin_plant_media', plant_id=plant.id)
    else:
        form = PlantMediaForm(instance=media)

    return render(request, 'store/admin/edit_plant_media.html', {
        'plant': plant,
        'media': media,
        'form': form,
    })


def add_to_cart_view(request, fish_id):
    fish = get_object_or_404(Fish, id=fish_id)
    try:
        quantity = int(request.POST.get('quantity', 1))
    except (TypeError, ValueError):
        quantity = 1

    # Check minimum order quantity
    if quantity < fish.minimum_order_quantity:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': f'Minimum order quantity for {fish.name} is {fish.minimum_order_quantity}.'}, status=400)
        messages.error(request, f'Minimum order quantity for {fish.name} is {fish.minimum_order_quantity}.')
        return redirect('fish_detail', fish_id=fish_id)

    if request.user.is_authenticated and getattr(request.user, 'role', None) == 'customer':
        cart_item, created = Cart.objects.get_or_create(
            user=request.user,
            fish=fish,
            defaults={'quantity': quantity}
        )

        if not created:
            cart_item.quantity += quantity
            cart_item.save()

        total_items = (
            Cart.objects.filter(user=request.user).aggregate(total=models.Sum('quantity'))['total'] or 0
        ) + (
            AccessoryCart.objects.filter(user=request.user).aggregate(total=models.Sum('quantity'))['total'] or 0
        ) + (
            PlantCart.objects.filter(user=request.user).aggregate(total=models.Sum('quantity'))['total'] or 0
        )
    else:
        _add_guest_fish(request, fish.id, quantity)
        total_items = _guest_cart_total_counter(request)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': f'{fish.name} added to cart.', 'total_items': int(total_items)})

    messages.success(request, f'{fish.name} added to cart.')
    return redirect('cart')


@require_POST
def add_combo_to_cart_view(request, combo_id):
    """Add all fishes from a combo to the current user's cart.

    Expects POST. Validates stock and minimums before adding.
    """
    combo = get_object_or_404(ComboOffer, id=combo_id, is_active=True)

    # Validate all items first
    errors = []
    for item in combo.items.select_related('fish').all():
        fish = item.fish
        qty = max(1, int(item.quantity or 1))
        if not fish.is_available or fish.stock_quantity <= 0:
            errors.append(f"{fish.name} is not available.")
        elif fish.stock_quantity < qty:
            errors.append(f"Not enough stock for {fish.name} (requested {qty}).")
        elif qty < fish.minimum_order_quantity:
            errors.append(f"Minimum order for {fish.name} is {fish.minimum_order_quantity}.")

    if errors:
        messages.error(request, 'Could not add combo to cart: ' + ' '.join(errors))
        return redirect('combos')

    # Add items to cart
    if request.user.is_authenticated and getattr(request.user, 'role', None) == 'customer':
        for item in combo.items.select_related('fish').all():
            fish = item.fish
            qty = max(1, int(item.quantity or 1))
            cart_item, created = Cart.objects.get_or_create(
                user=request.user,
                fish=fish,
                combo=combo,
                defaults={'quantity': qty}
            )
            if not created:
                cart_item.quantity += qty
                cart_item.save()
    else:
        for item in combo.items.select_related('fish').all():
            fish = item.fish
            qty = max(1, int(item.quantity or 1))
            _add_guest_fish(request, fish.id, qty, combo_id=combo.id)

    messages.success(request, f'Combo "{combo.title}" added to your cart.')
    return redirect('cart')


@require_POST
def add_accessory_to_cart_view(request, accessory_id):
    from .models import AccessoryCart, Accessory
    accessory = get_object_or_404(Accessory, id=accessory_id)

    # Safely parse quantity
    try:
        quantity = int(request.POST.get('quantity', 1))
    except (TypeError, ValueError):
        quantity = 1

    if quantity < 1:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Invalid quantity.'}, status=400)
        messages.error(request, 'Invalid quantity.')
        return redirect('accessory_detail', accessory_id=accessory_id)

    try:
        if request.user.is_authenticated and getattr(request.user, 'role', None) == 'customer':
            cart_item, created = AccessoryCart.objects.get_or_create(
                user=request.user,
                accessory=accessory,
                defaults={'quantity': quantity}
            )

            if not created:
                cart_item.quantity += quantity
                cart_item.save()

            total_items = (
                AccessoryCart.objects.filter(user=request.user).aggregate(total=models.Sum('quantity'))['total'] or 0
            ) + (
                Cart.objects.filter(user=request.user).aggregate(total=models.Sum('quantity'))['total'] or 0
            ) + (
                PlantCart.objects.filter(user=request.user).aggregate(total=models.Sum('quantity'))['total'] or 0
            )
        else:
            _add_guest_accessory(request, accessory.id, quantity)
            total_items = _guest_cart_total_counter(request)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'{accessory.name} added to cart.', 'total_items': int(total_items)})

        messages.success(request, f'{accessory.name} added to cart.')
        return redirect('cart')
    except Exception as e:
        print(f"Error adding accessory to cart: {e}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Server error adding to cart.'}, status=500)
        messages.error(request, 'Server error adding to cart.')
        return redirect('accessory_detail', accessory_id=accessory_id)


@require_POST
def add_plant_to_cart_view(request, plant_id):
    plant = get_object_or_404(Plant, id=plant_id, is_active=True)

    if plant.price is None:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'This plant is priced on request.'}, status=400)
        messages.error(request, 'This plant requires contacting support for pricing.')
        return redirect('plants')

    if plant.stock_quantity is not None and plant.stock_quantity <= 0:
        msg = f'{plant.name} is currently out of stock.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': msg}, status=400)
        messages.error(request, msg)
        return redirect('plants')

    try:
        quantity = int(request.POST.get('quantity', plant.minimum_order_quantity or 1))
    except (TypeError, ValueError):
        quantity = plant.minimum_order_quantity or 1

    quantity = max(quantity, 1)

    if quantity < plant.minimum_order_quantity:
        msg = f'Minimum order quantity for {plant.name} is {plant.minimum_order_quantity}.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': msg}, status=400)
        messages.error(request, msg)
        return redirect('plants')

    if plant.stock_quantity is not None and plant.stock_quantity > 0 and quantity > plant.stock_quantity:
        msg = f'Only {plant.stock_quantity} units available for {plant.name}.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': msg}, status=400)
        messages.error(request, msg)
        return redirect('plants')

    try:
        if request.user.is_authenticated and getattr(request.user, 'role', None) == 'customer':
            cart_item, created = PlantCart.objects.get_or_create(
                user=request.user,
                plant=plant,
                defaults={'quantity': quantity}
            )
            if not created:
                cart_item.quantity += quantity
                cart_item.save()

            total_items = (
                PlantCart.objects.filter(user=request.user).aggregate(total=models.Sum('quantity'))['total'] or 0
            ) + (
                Cart.objects.filter(user=request.user).aggregate(total=models.Sum('quantity'))['total'] or 0
            ) + (
                AccessoryCart.objects.filter(user=request.user).aggregate(total=models.Sum('quantity'))['total'] or 0
            )
        else:
            _add_guest_plant(request, plant.id, quantity)
            total_items = _guest_cart_total_counter(request)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'{plant.name} added to cart.', 'total_items': int(total_items)})

        messages.success(request, f'{plant.name} added to cart.')
        return redirect('cart')
    except Exception as exc:
        print(f"Error adding plant to cart: {exc}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Server error adding to cart.'}, status=500)
        messages.error(request, 'Server error adding to cart.')
        return redirect('plants')


def cart_view(request):
    guest_mode = not (request.user.is_authenticated and getattr(request.user, 'role', None) == 'customer')

    if not guest_mode:
        cart_items = Cart.objects.filter(user=request.user).select_related('fish', 'combo')
        accessory_items = AccessoryCart.objects.filter(user=request.user)
        plant_items = PlantCart.objects.filter(user=request.user).select_related('plant')
    else:
        cart_items, accessory_items, plant_items = _build_guest_cart_items(request)

    # Group cart items that belong to the same combo
    bundle_groups = []
    standalone_items = []
    combos_map = {}
    for item in cart_items:
        if item.combo_id:
            combos_map.setdefault(item.combo_id, {'combo': item.combo, 'items': []})['items'].append(item)
        else:
            standalone_items.append(item)

    # Convert map to list for template ordering
    for k, v in combos_map.items():
        combo = v.get('combo')
        # compute displayed bundle price (use combo.bundle_price if present)
        # Determine how many full bundles the user's cart contains for this combo.
        # For each fish in the combo, compute how many bundles it supports: floor(cart_qty / combo_item_qty)
        bundle_count = None
        per_bundle_value = 0
        if combo:
            combo_items = {ci.fish_id: ci.quantity for ci in combo.items.all()}
            # compute per-bundle value (sum of fish.price * qty) as fallback
            per_bundle_value = 0
            for ci in combo.items.select_related('fish').all():
                per_bundle_value += float(ci.fish.price) * int(ci.quantity)

            counts = []
            for it in v.get('items', []):
                req = combo_items.get(it.fish_id, 1)
                # avoid division by zero
                try:
                    counts.append(int(it.quantity) // int(req))
                except Exception:
                    counts.append(0)

            if counts:
                bundle_count = min(counts)
            else:
                bundle_count = 0

        # Compute display price: prefer combo.bundle_price * bundle_count, else per_bundle_value * bundle_count
        if combo and combo.bundle_price:
            v['display_price'] = float(combo.bundle_price) * (bundle_count or 0)
        else:
            v['display_price'] = (bundle_count or 0) * per_bundle_value

        v['bundle_count'] = bundle_count or 0
        bundle_groups.append(v)

    # Compute total considering bundle pricing when applicable.
    total = 0.0
    # Add standalone items
    total += sum(float(item.get_total()) for item in standalone_items)
    # Add bundles and leftovers
    for g in bundle_groups:
        combo = g.get('combo')
        items = g.get('items', [])
        bundle_count = g.get('bundle_count', 0)
        if combo and combo.bundle_price and bundle_count > 0:
            # Add bundle price for each full bundle
            total += float(combo.bundle_price) * bundle_count
            # compute leftovers per fish
            combo_items = {ci.fish_id: ci.quantity for ci in combo.items.all()}
            for it in items:
                req = combo_items.get(it.fish_id, 1)
                leftover = int(it.quantity) - (int(req) * bundle_count)
                if leftover > 0:
                    total += float(it.fish.price) * leftover
        else:
            # No bundle pricing; add full item totals
            total += sum(float(it.get_total()) for it in items)

    # Accessories
    total += sum(float(item.get_total()) for item in accessory_items)
    # Plants
    total += sum(float(item.get_total()) for item in plant_items)

    return render(request, 'store/customer/cart.html', {
        'bundle_groups': bundle_groups,
        'cart_items': standalone_items,
        'accessory_items': accessory_items,
        'plant_items': plant_items,
        'total': total,
        'guest_mode': guest_mode,
    })


@login_required
@user_passes_test(is_customer)
def update_cart_view(request, cart_id):
    cart_item = get_object_or_404(Cart, id=cart_id, user=request.user)
    quantity = int(request.POST.get('quantity', 1))
    
    # Check minimum order quantity
    if quantity > 0 and quantity < cart_item.fish.minimum_order_quantity:
        messages.error(request, f'Minimum order quantity for {cart_item.fish.name} is {cart_item.fish.minimum_order_quantity}.')
        return redirect('cart')
    
    if quantity > 0:
        cart_item.quantity = quantity
        cart_item.save()
    else:
        cart_item.delete()
    
    return redirect('cart')


@login_required
@user_passes_test(is_customer)
def remove_from_cart_view(request, cart_id):
    cart_item = get_object_or_404(Cart, id=cart_id, user=request.user)
    cart_item.delete()
    messages.success(request, 'Item removed from cart.')
    return redirect('cart')


@login_required
@user_passes_test(is_customer)
def remove_bundle_view(request, combo_id):
    """Remove all cart items associated with a combo for the current user."""
    Cart.objects.filter(user=request.user, combo_id=combo_id).delete()
    messages.success(request, 'Bundle removed from cart.')
    return redirect('cart')


@login_required
@user_passes_test(is_customer)
def update_accessory_cart_view(request, accessory_cart_id):
    a_item = get_object_or_404(AccessoryCart, id=accessory_cart_id, user=request.user)
    try:
        quantity = int(request.POST.get('quantity', 1))
    except (TypeError, ValueError):
        quantity = 1

    # Check minimum order quantity
    if quantity > 0 and quantity < a_item.accessory.minimum_order_quantity:
        messages.error(request, f'Minimum order quantity for {a_item.accessory.name} is {a_item.accessory.minimum_order_quantity}.')
        return redirect('cart')

    if quantity > 0:
        a_item.quantity = quantity
        a_item.save()
    else:
        a_item.delete()

    return redirect('cart')


@login_required
@user_passes_test(is_customer)
def remove_accessory_cart_view(request, accessory_cart_id):
    a_item = get_object_or_404(AccessoryCart, id=accessory_cart_id, user=request.user)
    a_item.delete()
    messages.success(request, 'Accessory removed from cart.')
    return redirect('cart')


@login_required
@user_passes_test(is_customer)
def update_plant_cart_view(request, plant_cart_id):
    p_item = get_object_or_404(PlantCart, id=plant_cart_id, user=request.user)
    try:
        quantity = int(request.POST.get('quantity', 1))
    except (TypeError, ValueError):
        quantity = p_item.quantity

    if quantity > 0 and quantity < p_item.plant.minimum_order_quantity:
        messages.error(request, f'Minimum order quantity for {p_item.plant.name} is {p_item.plant.minimum_order_quantity}.')
        return redirect('cart')

    available = p_item.plant.stock_quantity
    if available is not None and available >= 0 and quantity > available:
        messages.error(request, f'Only {available} units of {p_item.plant.name} available.')
        return redirect('cart')

    if quantity > 0:
        p_item.quantity = quantity
        p_item.save()
    else:
        p_item.delete()

    return redirect('cart')


@login_required
@user_passes_test(is_customer)
def remove_plant_cart_view(request, plant_cart_id):
    p_item = get_object_or_404(PlantCart, id=plant_cart_id, user=request.user)
    p_item.delete()
    messages.success(request, 'Plant removed from cart.')
    return redirect('cart')




@require_POST
def apply_coupon_view(request):
    """AJAX view to apply coupon code. Allows anonymous users to save a coupon
    code in session for later application (guest flow). If the cart is empty
    for anonymous users the coupon will be saved and a friendly message is
    returned."""
    try:
        # Allow anonymous users to apply a coupon code (store in session).
        # Frontend will later use this to show the coupon when/if a cart exists.

        coupon_code = request.POST.get('coupon_code', '').strip().upper()
        preview_flag = str(request.POST.get('preview', '')).lower() in ('1', 'true', 'yes')

        if not coupon_code:
            return JsonResponse({'success': False, 'message': 'Please enter a coupon code'})

        try:
            coupon = Coupon.objects.get(code=coupon_code)
        except Coupon.DoesNotExist:
            # Allow anonymous users to save arbitrary coupon codes into session
            # (so they can paste a code before registering/logging-in). For
            # authenticated users we still require a valid coupon.
            if not request.user.is_authenticated:
                if not preview_flag:
                    request.session['applied_coupon_code'] = coupon_code
                    return JsonResponse({
                        'success': True,
                        'message': 'Coupon saved. It will be applied when valid.',
                        'discount': 0.0,
                        'final_total': 0.0,
                        'coupon_code': coupon_code,
                    })
                else:
                    # Preview for unknown coupon: report invalid for preview
                    return JsonResponse({'success': False, 'message': 'Invalid coupon code'})
            return JsonResponse({'success': False, 'message': 'Invalid coupon code'})

        # If not force_apply, validate coupon normally
        now = timezone.now()
        if not getattr(coupon, 'force_apply', False):
            if not coupon.is_active:
                return JsonResponse({'success': False, 'message': 'This coupon is no longer active'})

            # Check validity with detailed error
            if coupon.valid_from and now < coupon.valid_from:
                return JsonResponse({
                    'success': False,
                    'message': f'This coupon is not valid yet. Valid from: {coupon.valid_from.strftime("%d %b %Y, %I:%M %p")}'
                })

            if coupon.valid_until and now > coupon.valid_until:
                return JsonResponse({
                    'success': False,
                    'message': 'This coupon has expired'
                })

            # Check usage limit
            if coupon.usage_limit and coupon.times_used >= coupon.usage_limit:
                return JsonResponse({'success': False, 'message': 'This coupon has reached its usage limit'})

            # Check if user can use this coupon (treat anonymous as not favorite)
            is_fav = getattr(request.user, 'is_favorite', False)
            if coupon.coupon_type == 'favorites' and not is_fav:
                return JsonResponse({'success': False, 'message': 'This coupon is only for favorite customers'})

            if coupon.coupon_type == 'normal' and is_fav:
                return JsonResponse({'success': False, 'message': 'This coupon is only for normal users'})

        # Calculate cart total including accessories. For anonymous users the
        # DB-backed cart will be empty; total will be 0. We still allow saving
        # the coupon into session so it can be applied later when a cart exists.
        cart_items = Cart.objects.filter(user=request.user).select_related('fish', 'combo')
        accessory_items = AccessoryCart.objects.filter(user=request.user).select_related('accessory')
        plant_items = PlantCart.objects.filter(user=request.user).select_related('plant')
        total = (
            sum(item.get_total() for item in cart_items)
            + sum(item.get_total() for item in accessory_items)
            + sum(item.get_total() for item in plant_items)
        )

        total_weight = _calculate_total_weight(cart_items, accessory_items, plant_items)

        shipping_state = request.POST.get('shipping_state') or getattr(request.user, 'address', '')
        shipping_pincode = request.POST.get('shipping_pincode') or ''
        shipping_address = request.POST.get('shipping_address') or getattr(request.user, 'address', '')
        try:
            delivery_charge, delivery_rate = _calculate_delivery_charge(
                total_weight,
                state=shipping_state,
                pincode=shipping_pincode,
                address=shipping_address,
            )
        except ShippingUnavailableError as exc:
            return JsonResponse({
                'success': False,
                'message': f'Delivery is not available in {exc.state}. Please update your shipping state.'
            }, status=400)
        kerala_rate_value, default_rate_value, _ = _get_shipping_rates()

        # Check minimum order amount unless force_apply
        if not getattr(coupon, 'force_apply', False):
            if coupon.min_order_amount and total < coupon.min_order_amount:
                return JsonResponse({
                    'success': False,
                    'message': f'Minimum order amount of ₹{coupon.min_order_amount} required'
                })

        # Calculate discount
        discount = (total * coupon.discount_percentage) / 100
        if coupon.max_discount_amount:
            discount = min(discount, coupon.max_discount_amount)

        final_total = max(Decimal('0'), Decimal(total) - Decimal(discount)) + delivery_charge


        # Store in session so the coupon persists for the visitor (unless preview)
        if not preview_flag:
            request.session['applied_coupon_code'] = coupon_code

        # If there's no cart total (guest without DB cart), inform the user
        # that the coupon is saved for later use. Otherwise return the actual
        # discount and final total.
        if total == 0:
            return JsonResponse({
                'success': True,
                'message': 'Coupon saved. It will be applied when you have items in your cart.' if not preview_flag else 'Coupon preview: no items in cart',
                'discount': 0.0,
                'final_total': 0.0,
                'coupon_code': coupon_code,
                'discount_percentage': float(coupon.discount_percentage),
                'delivery_charge': float(delivery_charge),
                'delivery_rate': float(delivery_rate),
                'total_weight': float(total_weight),
                'kerala_delivery_rate': float(kerala_rate_value),
                'default_delivery_rate': float(default_rate_value),
                'preview': preview_flag,
            })

        return JsonResponse({
            'success': True,
            'message': f'Coupon applied! You saved ₹{discount:.2f}' if not preview_flag else f'Coupon preview: you would save ₹{discount:.2f}',
            'discount': float(discount),
            'final_total': float(final_total),
            'coupon_code': coupon_code,
            'discount_percentage': float(coupon.discount_percentage),
            'delivery_charge': float(delivery_charge),
            'delivery_rate': float(delivery_rate),
            'total_weight': float(total_weight),
            'kerala_delivery_rate': float(kerala_rate_value),
            'default_delivery_rate': float(default_rate_value),
            'preview': preview_flag,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@require_POST
def remove_coupon_view(request):
    """AJAX view to remove applied coupon"""
    if 'applied_coupon_code' in request.session:
        del request.session['applied_coupon_code']
    
    cart_items = Cart.objects.filter(user=request.user).select_related('fish', 'combo')
    accessory_items = AccessoryCart.objects.filter(user=request.user).select_related('accessory')
    plant_items = PlantCart.objects.filter(user=request.user).select_related('plant')
    total = (
        sum(item.get_total() for item in cart_items)
        + sum(item.get_total() for item in accessory_items)
        + sum(item.get_total() for item in plant_items)
    )
    total_weight = _calculate_total_weight(cart_items, accessory_items, plant_items)

    shipping_state = request.POST.get('shipping_state') or getattr(request.user, 'address', '')
    shipping_pincode = request.POST.get('shipping_pincode') or ''
    shipping_address = request.POST.get('shipping_address') or getattr(request.user, 'address', '')
    try:
        delivery_charge, delivery_rate = _calculate_delivery_charge(
            total_weight,
            state=shipping_state,
            pincode=shipping_pincode,
            address=shipping_address,
        )
    except ShippingUnavailableError as exc:
        return JsonResponse({
            'success': False,
            'message': f'Delivery is not available in {exc.state}. Please update your shipping state.'
        }, status=400)
    kerala_rate_value, default_rate_value, _ = _get_shipping_rates()

    final_total = max(Decimal('0'), Decimal(total)) + delivery_charge
    # When removed, discount is zero and final_total equals total
    return JsonResponse({
        'success': True,
        'message': 'Coupon removed',
        'total': float(total),
        'final_total': float(final_total),
        'discount': 0.0,
        'coupon_code': '',
        'discount_percentage': 0.0,
        'delivery_charge': float(delivery_charge),
        'delivery_rate': float(delivery_rate),
        'total_weight': float(total_weight),
        'kerala_delivery_rate': float(kerala_rate_value),
        'default_delivery_rate': float(default_rate_value),
    })


def _build_upi_metadata(order):
    """Return reusable UPI payment metadata for an order (QR, deeplinks)."""
    upi_id = getattr(settings, 'UPI_PAYMENT_ID', 'muhzinmuhammed4@oksbi')
    merchant_name = getattr(settings, 'SITE_NAME', 'Fishy Friend Aquatics')
    amount = str(order.final_amount or order.total_amount or '0')
    transaction_note = f"Order {order.order_number}"

    upi_string = f"upi://pay?pa={upi_id}&pn={merchant_name}&am={amount}&tn={transaction_note}&cu=INR"

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(upi_string)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()

    upi_apps = {
        'phonepe': f"phonepe://pay?pa={upi_id}&pn={merchant_name}&am={amount}&tn={transaction_note}&cu=INR",
        'googlepay': f"gpay://upi/pay?pa={upi_id}&pn={merchant_name}&am={amount}&tn={transaction_note}&cu=INR",
        'paytm': f"paytmmp://pay?pa={upi_id}&pn={merchant_name}&am={amount}&tn={transaction_note}&cu=INR",
    }

    return {
        'upi_id': upi_id,
        'upi_string': upi_string,
        'qr_code': qr_code_base64,
        'upi_apps': upi_apps,
    }


def create_draft_order(request):
    """AJAX endpoint: create or return a recent draft Order for the current user's cart.

    Returns JSON: { order_id, order_number, final_amount }
    """
    try:
        cart_items = Cart.objects.filter(user=request.user).select_related('fish', 'combo')
        accessory_items = AccessoryCart.objects.filter(user=request.user).select_related('accessory')
        plant_items = PlantCart.objects.filter(user=request.user).select_related('plant')

        if not cart_items and not accessory_items and not plant_items:
            return JsonResponse({'error': 'Cart is empty'}, status=400)

        total = (
            sum(item.get_total() for item in cart_items)
            + sum(item.get_total() for item in accessory_items)
            + sum(item.get_total() for item in plant_items)
        )
        total = Decimal(total)
        total_weight = _calculate_total_weight(cart_items, accessory_items, plant_items)

        # Get applied coupon from session if present
        applied_coupon = None
        discount = Decimal('0')
        final_total = total
        if 'applied_coupon_code' in request.session:
            try:
                coupon = Coupon.objects.get(code=request.session['applied_coupon_code'])
                if coupon.is_valid() and coupon.can_use(request.user):
                    applied_coupon = coupon
                    discount = (total * Decimal(coupon.discount_percentage)) / Decimal('100')
                    if coupon.max_discount_amount:
                        discount = min(discount, Decimal(coupon.max_discount_amount))
                    final_total = total - discount
            except Coupon.DoesNotExist:
                pass

        shipping_address = (request.POST.get('shipping_address')
                             or getattr(request.user, 'address', '')
                             or '')
        shipping_state = request.POST.get('shipping_state', '')
        shipping_pincode = request.POST.get('shipping_pincode', '')
        phone_number = (request.POST.get('phone_number')
                        or getattr(request.user, 'phone_number', '')
                        or '')
        payment_method = request.POST.get('payment_method') or 'card'

        try:
            delivery_charge, delivery_rate = _calculate_delivery_charge(
                total_weight,
                state=shipping_state,
                pincode=shipping_pincode,
                address=shipping_address,
            )
        except ShippingUnavailableError as exc:
            return JsonResponse({
                'error': f'Delivery is not available in {exc.state}. Please choose a different state.',
                'shipping_blocked': True
            }, status=400)
        kerala_rate_value, default_rate_value, _ = _get_shipping_rates()
        final_total = max(Decimal('0'), final_total) + delivery_charge

        # Try to reuse a recent draft
        draft_cutoff = timezone.now() - timedelta(minutes=30)
        draft_order = Order.objects.filter(
            user=request.user,
            transaction_id__isnull=True,
            status='pending',
            created_at__gte=draft_cutoff,
        ).order_by('-created_at').first()

        if draft_order is None:
            draft_order = Order.objects.create(
                user=request.user,
                order_number=Order.generate_order_number(),
                total_amount=total,
                coupon=applied_coupon,
                discount_amount=discount,
                final_amount=final_total,
                delivery_charge=delivery_charge,
                total_weight=total_weight,
                shipping_address=shipping_address,
                shipping_state=shipping_state,
                shipping_pincode=shipping_pincode,
                phone_number=phone_number,
                payment_method=payment_method,
                payment_status='pending',
            )
            for cart_item in cart_items:
                OrderItem.objects.create(
                    order=draft_order,
                    fish=cart_item.fish,
                    quantity=cart_item.quantity,
                    price=cart_item.fish.price,
                )
            for a_item in accessory_items:
                OrderAccessoryItem.objects.create(
                    order=draft_order,
                    accessory=a_item.accessory,
                    quantity=a_item.quantity,
                    price=a_item.accessory.price,
                )
            for p_item in plant_items:
                OrderPlantItem.objects.create(
                    order=draft_order,
                    plant=p_item.plant,
                    quantity=p_item.quantity,
                    price=p_item.plant.price,
                )
        else:
            # Refresh core order fields to match the current cart snapshot
            draft_order.total_amount = total
            draft_order.coupon = applied_coupon
            draft_order.discount_amount = discount
            draft_order.final_amount = final_total
            draft_order.shipping_address = shipping_address
            draft_order.shipping_state = shipping_state
            draft_order.shipping_pincode = shipping_pincode
            draft_order.phone_number = phone_number
            draft_order.payment_method = payment_method
            draft_order.status = 'pending'
            draft_order.payment_status = 'pending'
            draft_order.transaction_id = None
            draft_order.provider_order_id = None
            draft_order.delivery_charge = delivery_charge
            draft_order.total_weight = total_weight
            draft_order.save()

            # Replace line items so the order mirrors the latest cart contents
            draft_order.items.all().delete()
            for cart_item in cart_items:
                OrderItem.objects.create(
                    order=draft_order,
                    fish=cart_item.fish,
                    quantity=cart_item.quantity,
                    price=cart_item.fish.price,
                )

            draft_order.accessory_items.all().delete()
            for a_item in accessory_items:
                OrderAccessoryItem.objects.create(
                    order=draft_order,
                    accessory=a_item.accessory,
                    quantity=a_item.quantity,
                    price=a_item.accessory.price,
                )
            draft_order.plant_items.all().delete()
            for p_item in plant_items:
                OrderPlantItem.objects.create(
                    order=draft_order,
                    plant=p_item.plant,
                    quantity=p_item.quantity,
                    price=p_item.plant.price,
                )

        response_data = {
            'order_id': draft_order.id,
            'order_number': draft_order.order_number,
            'final_amount': float(draft_order.final_amount),
            'payment_method': draft_order.payment_method,
            'delivery_charge': float(delivery_charge),
            'delivery_rate': float(delivery_rate),
            'total_weight': float(total_weight),
            'kerala_delivery_rate': float(kerala_rate_value),
            'default_delivery_rate': float(default_rate_value),
        }

        if draft_order.payment_method == 'upi':
            try:
                upi_meta = _build_upi_metadata(draft_order)
                response_data.update({
                    'upi_qr': upi_meta['qr_code'],
                    'upi_id': upi_meta['upi_id'],
                    'upi_string': upi_meta['upi_string'],
                    'upi_apps': upi_meta['upi_apps'],
                })
            except Exception:
                logging.getLogger(__name__).exception('Failed generating UPI metadata for order %s', draft_order.order_number)

        return JsonResponse(response_data)
    except Exception:
        logging.getLogger(__name__).exception('Failed to create draft order via AJAX')
        return JsonResponse({'error': 'Server error'}, status=500)


@login_required
@user_passes_test(is_customer)
def customer_orders_view(request):
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'store/customer/orders.html', {'orders': orders})


@login_required
@user_passes_test(is_customer)
def order_detail_view(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    existing_review = Review.objects.filter(user=request.user, order=order).first()
    review_form = None
    if order.status == 'delivered' and not existing_review:
        review_form = ReviewForm()
    return render(request, 'store/customer/order_detail.html', {
        'order': order,
        'review_form': review_form,
        'existing_review': existing_review,
    })


@login_required
@user_passes_test(is_customer)
def order_confirmation_view(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    invoice_url = None
    if order.payment_status == 'paid':
        invoice_filename = f'invoice-{order.order_number}.pdf'
        invoice_path = os.path.join(settings.MEDIA_ROOT, 'invoices', invoice_filename)
        if os.path.exists(invoice_path):
            invoice_url = settings.MEDIA_URL + f'invoices/{invoice_filename}'
    return render(request, 'store/customer/order_confirmation.html', {
        'order': order,
        'invoice_url': invoice_url,
    })


@login_required
@user_passes_test(is_customer)
def resume_payment_view(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)

    if order.payment_status == 'paid':
        messages.info(request, 'This order is already paid. Showing your latest confirmation.')
        return redirect('order_confirmation', order_id=order.id)

    if order.status == 'cancelled':
        messages.error(request, 'Cannot process payment for a cancelled order.')
        return redirect('order_detail', order_id=order.id)

    if order.payment_method == 'upi':
        return redirect('upi_payment', order_id=order.id)

    return render(request, 'store/customer/resume_payment.html', {
        'order': order,
    })


@login_required
@user_passes_test(is_customer)
def submit_review_view(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    if order.status != 'delivered':
        messages.error(request, 'You can only review a delivered order.')
        return redirect('order_detail', order_id=order.id)
    if Review.objects.filter(user=request.user, order=order).exists():
        messages.info(request, 'You have already reviewed this order.')
        return redirect('order_detail', order_id=order.id)
    form = ReviewForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            review = form.save(commit=False)
            review.user = request.user
            review.order = order
            review.save()
            messages.success(request, 'Review submitted successfully!')
            return redirect('order_detail', order_id=order.id)
        else:
            messages.error(request, 'Please correct the errors in your review form.')
    return redirect('order_detail', order_id=order.id)


@login_required
@user_passes_test(is_customer)
def cancel_order_view(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    if request.method == 'POST':
        # Only allow cancellation for pending or processing orders
        if order.status in ['pending', 'processing']:
            order.status = 'cancelled'
            order.save()
            # Send cancellation email to customer
            try:
                subject = f'Order Cancelled - {settings.SITE_NAME} - {order.order_number}'
                try:
                    from store.tasks import send_order_email
                    site_base = request.build_absolute_uri('/') if request is not None else getattr(settings, 'SITE_URL', None)
                    send_order_email.delay(order.id, 'order_cancelled', subject, order.user.email, site_base)
                except Exception:
                    logging.getLogger(__name__).exception('Celery task import failed; falling back to synchronous send for order %s', order.order_number)
                    _send_order_email(order, 'order_cancelled', subject, order.user.email, request=request)
            except Exception:
                logging.getLogger(__name__).exception('Error sending cancellation email for order %s', order.order_number)
            messages.success(request, f'Order #{order.order_number} has been cancelled successfully.')
        else:
            messages.error(request, f'Order #{order.order_number} cannot be cancelled at this stage.')
        
        return redirect('order_detail', order_id=order.id)
    
    return redirect('order_detail', order_id=order.id)


@login_required
@user_passes_test(is_admin)
def admin_cancel_order_view(request, order_id):
    """Admin endpoint to cancel an order and redirect to admin orders list."""
    order = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        # Only allow cancellation for pending or processing orders
        if order.status in ['pending', 'processing']:
            order.status = 'cancelled'
            order.save()
            # Send cancellation email to customer
            try:
                subject = f'Order Cancelled - {settings.SITE_NAME} - {order.order_number}'
                try:
                    from store.tasks import send_order_email
                    site_base = request.build_absolute_uri('/') if request is not None else getattr(settings, 'SITE_URL', None)
                    send_order_email.delay(order.id, 'order_cancelled', subject, order.user.email, site_base)
                except Exception:
                    logging.getLogger(__name__).exception('Celery task import failed; falling back to synchronous send for order %s', order.order_number)
                    _send_order_email(order, 'order_cancelled', subject, order.user.email, request=request)
            except Exception:
                logging.getLogger(__name__).exception('Error sending cancellation email for order %s', order.order_number)
            messages.success(request, f'Order #{order.order_number} has been cancelled successfully.')
        else:
            messages.error(request, f'Order #{order.order_number} cannot be cancelled at this stage.')
    
    return redirect('admin_orders')


# UPI Payment Views
@login_required
@user_passes_test(is_customer)
def upi_payment_view(request, order_id):
    """Display UPI payment page with QR code and payment options"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    # Check if payment is already completed
    if order.payment_status == 'paid':
        messages.info(request, 'This order has already been paid.')

    # Check if order is cancelled
    if order.status == 'cancelled':
        messages.error(request, 'Cannot process payment for a cancelled order.')
        return redirect('order_detail', order_id=order.id)
    
    # Generate UPI payment string
    # Format: upi://pay?pa=UPI_ID&pn=MERCHANT_NAME&am=AMOUNT&tn=TRANSACTION_NOTE&cu=INR
    upi_meta = _build_upi_metadata(order)

    context = {
        'order': order,
        'upi_id': upi_meta['upi_id'],
        'upi_string': upi_meta['upi_string'],
        'qr_code': upi_meta['qr_code'],
        'upi_apps': upi_meta['upi_apps'],
    }
    
    return render(request, 'store/customer/upi_payment.html', context)


@login_required
@user_passes_test(is_staff_user)
@require_http_methods(['GET', 'POST'])
def admin_add_combo_view(request):
    """Staff-only page: create a ComboOffer with multiple fish items.

    Form fields (simple): title, bundle_price (optional), is_active
    Items: multiple rows of fish_id[] and quantity[] submitted via JS
    """
    combo_categories = Category.objects.filter(category_type='combo').order_by('name')
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        raw_bundle_price = (request.POST.get('bundle_price') or '').strip()
        raw_weight = (request.POST.get('weight') or '').strip()
        bundle_price = None
        weight = None
        is_active = request.POST.get('is_active') == 'on'
        show_on_homepage = request.POST.get('show_on_homepage') == 'on'
        category_id = request.POST.get('category_id') or None
        fish_ids = request.POST.getlist('fish_id')
        quantities = request.POST.getlist('quantity')
        errors = []
        if not title:
            errors.append('Title is required.')
        if not fish_ids:
            errors.append('Add at least one fish to the combo.')

        selected_category = None
        if category_id:
            try:
                selected_category = combo_categories.get(id=int(category_id))
            except (ValueError, Category.DoesNotExist):
                selected_category = None
        if not selected_category:
            errors.append('Select a category for this combo.')

        if raw_bundle_price:
            try:
                bundle_price = Decimal(raw_bundle_price)
                if bundle_price < 0:
                    errors.append('Bundle price cannot be negative.')
            except InvalidOperation:
                errors.append('Bundle price must be a valid decimal number.')

        if raw_weight:
            try:
                weight = Decimal(raw_weight)
                if weight < 0:
                    errors.append('Weight cannot be negative.')
            except InvalidOperation:
                errors.append('Weight must be a valid decimal number.')

        if errors:
            fishes = Fish.objects.all().order_by('name')
            combos = ComboOffer.objects.all().order_by('-created_at')
            return render(request, 'store/admin/add_combo.html', {
                'errors': errors,
                'fishes': fishes,
                'combos': combos,
                'categories': combo_categories,
                'title': title,
                'bundle_price': raw_bundle_price,
                'weight': raw_weight,
                'is_active': is_active,
                'show_on_homepage': show_on_homepage,
                'selected_category_id': category_id,
            })

        combo = ComboOffer.objects.create(
            title=title,
            bundle_price=bundle_price,
            weight=weight,
            is_active=is_active,
            show_on_homepage=show_on_homepage,
            category=selected_category,
        )
        for idx, fid in enumerate(fish_ids):
            try:
                fish = Fish.objects.get(id=int(fid))
                qty = int(quantities[idx]) if idx < len(quantities) else 1
                ComboItem.objects.create(combo=combo, fish=fish, quantity=max(1, qty))
            except Exception:
                continue

        messages.success(request, f'Combo "{combo.title}" created successfully.')
        return redirect('admin_fishes')

    # GET
    fishes = Fish.objects.all().order_by('name')
    # also include existing combos so staff can edit from the same page
    combos = ComboOffer.objects.all().order_by('-created_at')

    # Support pre-filling the form from staff links, e.g. ?prefill_fish=12 or ?prefill_fish=12,34
    prefill_param = request.GET.get('prefill_fish') or request.GET.get('preselect_fish')
    prefill_fish_ids = []
    if prefill_param:
        try:
            # allow comma separated values
            parts = [p.strip() for p in str(prefill_param).split(',') if p.strip()]
            for p in parts:
                fid = int(p)
                # ensure fish exists
                if Fish.objects.filter(id=fid).exists():
                    prefill_fish_ids.append(fid)
        except Exception:
            prefill_fish_ids = []

    return render(request, 'store/admin/add_combo.html', {
        'fishes': fishes,
        'combos': combos,
        'prefill_fish_ids': prefill_fish_ids,
        'categories': combo_categories,
        'selected_category_id': '',
        'is_active': True,
        'show_on_homepage': False,
    })
    
    # Check if order is cancelled
    if order.status == 'cancelled':
        messages.error(request, 'Cannot process payment for a cancelled order.')


@login_required
@user_passes_test(is_staff_user)
@require_http_methods(['GET', 'POST'])
def admin_edit_combo_view(request, combo_id):
    """Edit an existing combo. Replaces items with submitted list."""
    combo = get_object_or_404(ComboOffer, id=combo_id)
    combo_categories = Category.objects.filter(category_type='combo').order_by('name')
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        raw_bundle_price = (request.POST.get('bundle_price') or '').strip()
        raw_weight = (request.POST.get('weight') or '').strip()
        bundle_price = None
        weight = None
        is_active = request.POST.get('is_active') == 'on'
        show_on_homepage = request.POST.get('show_on_homepage') == 'on'
        category_id = request.POST.get('category_id') or None
        fish_ids = request.POST.getlist('fish_id')
        quantities = request.POST.getlist('quantity')
        errors = []
        if not title:
            errors.append('Title is required.')
        if not fish_ids:
            errors.append('Add at least one fish to the combo.')

        selected_category = None
        if category_id:
            try:
                selected_category = combo_categories.get(id=int(category_id))
            except (ValueError, Category.DoesNotExist):
                selected_category = None
        if not selected_category:
            errors.append('Select a category for this combo.')

        if raw_bundle_price:
            try:
                bundle_price = Decimal(raw_bundle_price)
                if bundle_price < 0:
                    errors.append('Bundle price cannot be negative.')
            except InvalidOperation:
                errors.append('Bundle price must be a valid decimal number.')

        if raw_weight:
            try:
                weight = Decimal(raw_weight)
                if weight < 0:
                    errors.append('Weight cannot be negative.')
            except InvalidOperation:
                errors.append('Weight must be a valid decimal number.')

        if errors:
            fishes = Fish.objects.all().order_by('name')
            combos = ComboOffer.objects.all().order_by('-created_at')
            return render(request, 'store/admin/add_combo.html', {
                'errors': errors,
                'fishes': fishes,
                'combos': combos,
                'editing': True,
                'combo': combo,
                'categories': combo_categories,
                'selected_category_id': category_id,
                'title': title,
                'bundle_price': raw_bundle_price,
                'weight': raw_weight,
                'is_active': is_active,
                'show_on_homepage': show_on_homepage,
            })

        # Update combo fields
        combo.title = title
        combo.bundle_price = bundle_price
        combo.weight = weight
        combo.is_active = is_active
        combo.show_on_homepage = show_on_homepage
        combo.category = selected_category
        combo.save()

        # Replace items: simple approach - delete existing and recreate
        ComboItem.objects.filter(combo=combo).delete()
        for idx, fid in enumerate(fish_ids):
            try:
                fish = Fish.objects.get(id=int(fid))
                qty = int(quantities[idx]) if idx < len(quantities) else 1
                ComboItem.objects.create(combo=combo, fish=fish, quantity=max(1, qty))
            except Exception:
                continue

        messages.success(request, f'Combo "{combo.title}" updated successfully.')
        return redirect('admin_fishes')

    # GET - render form prefilled
    fishes = Fish.objects.all().order_by('name')
    combos = ComboOffer.objects.all().order_by('-created_at')
    return render(request, 'store/admin/add_combo.html', {
        'fishes': fishes,
        'combos': combos,
        'editing': True,
        'combo': combo,
        'categories': combo_categories,
        'selected_category_id': str(combo.category_id) if combo.category_id else '',
        'is_active': combo.is_active,
        'show_on_homepage': combo.show_on_homepage,
    })


@login_required
@user_passes_test(is_staff_user)
@require_POST
def admin_delete_combo_view(request, combo_id):
    """Delete a combo. Accessible by staff only. Expects POST (CSRF-protected)."""
    combo = get_object_or_404(ComboOffer, id=combo_id)
    title = combo.title
    try:
        combo.delete()
        messages.success(request, f'Combo "{title}" deleted successfully.')
    except Exception:
        messages.error(request, f'Unable to delete combo "{title}". Please try again.')
    return redirect('admin_add_combo')


@login_required
@user_passes_test(is_customer)
def verify_upi_payment(request, order_id):
    """Verify UPI payment and update order status"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    if request.method == 'POST':
        # In a real scenario, you would verify with payment gateway
        # For now, we'll simulate successful payment
        upi_transaction_id = request.POST.get('upi_transaction_id', '').strip()
        
        if upi_transaction_id:
            try:
                processed, order = finalize_order_payment(order, payment_id=upi_transaction_id, request=request)
            except Exception:
                logging.getLogger(__name__).exception('Failed to finalize UPI payment for order %s', order.order_number)
                messages.error(request, 'We could not verify the payment right now. Please contact support with your transaction ID.')
                return redirect('upi_payment', order_id=order.id)

            if processed:
                messages.success(request, 'Payment successful! Your order has been confirmed and an invoice has been sent.')
            else:
                messages.info(request, 'Payment was already verified earlier. Showing your confirmation details.')
            return redirect('order_confirmation', order_id=order.id)
        else:
            messages.error(request, 'Please enter a valid UPI transaction ID.')
            return redirect('upi_payment', order_id=order.id)
    
    return redirect('upi_payment', order_id=order.id)


# Staff Views
@login_required
@user_passes_test(is_staff)
def staff_dashboard_view(request):
    total_fishes = Fish.objects.count()
    total_categories = Category.objects.count()
    total_breeds = Breed.objects.count()
    total_orders = Order.objects.count()
    
    return render(request, 'store/staff/dashboard.html', {
        'total_fishes': total_fishes,
        'total_categories': total_categories,
        'total_breeds': total_breeds,
        'total_orders': total_orders,
    })


@login_required
@user_passes_test(is_staff)
def staff_fish_list_view(request):
    fishes = Fish.objects.select_related('category', 'breed').all()

    search_query = (request.GET.get('search') or '').strip()
    category_id = (request.GET.get('category') or '').strip()
    breed_id = (request.GET.get('breed') or '').strip()
    availability = (request.GET.get('availability') or '').strip()
    featured = (request.GET.get('featured') or '').strip()

    if search_query:
        fishes = fishes.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if category_id:
        fishes = fishes.filter(category_id=category_id)

    if breed_id:
        fishes = fishes.filter(breed_id=breed_id)


    if availability == 'available':
        fishes = fishes.filter(is_available=True)
    elif availability == 'unavailable':
        fishes = fishes.filter(is_available=False)
    elif availability == 'in_stock':
        fishes = fishes.filter(stock_quantity__gt=0)
    elif availability == 'out_of_stock':
        fishes = fishes.filter(stock_quantity__lte=0)

    if featured == 'yes':
        fishes = fishes.filter(is_featured=True)
    elif featured == 'no':
        fishes = fishes.filter(is_featured=False)
    # If AJAX, return only the rendered table HTML so the client can update it
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render_to_string('store/staff/_fish_table.html', {'fishes': fishes}, request=request)
        # If there are no fishes, still return the empty-state HTML block so the client can replace container
        if not fishes:
            empty_html = render_to_string('store/staff/_fish_empty.html', request=request)
            return JsonResponse({'html': html + empty_html})
        return JsonResponse({'html': html})

    categories_with_count = Category.objects.filter(category_type='fish').annotate(
        fish_count=Count('fishes', distinct=True)
    ).order_by('name')
    
    return render(request, 'store/staff/fish_list.html', {
        'fishes': fishes,
        'search_query': search_query,
        'categories': categories_with_count,
        'breeds': Breed.objects.all().order_by('name'),
        'category_id': category_id,
        'breed_id': breed_id,
        'availability': availability,
        'featured': featured,
    })


@login_required
@user_passes_test(is_staff)
def add_category_view(request):
    type_hint = request.GET.get('type') or request.POST.get('category_type')
    valid_types = dict(Category.CATEGORY_TYPES)
    if type_hint not in valid_types:
        type_hint = None

    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category added successfully.')
            return redirect('staff_categories')
    else:
        initial = {}
        if type_hint:
            initial['category_type'] = type_hint
        form = CategoryForm(initial=initial)

    type_label = valid_types.get(type_hint) if type_hint else None

    return render(
        request,
        'store/staff/add_category.html',
        {
            'form': form,
            'type_hint': type_hint,
            'type_label': type_label,
        },
    )


@login_required
@user_passes_test(is_staff)
def staff_categories_view(request):
    type_configs = {
        'fish': {
            'title': 'Fish Categories',
            'icon': 'fas fa-fish',
            'badge': 'Fish',
            'add_label': 'Add Fish Category',
        },
        'combo': {
            'title': 'Combo Offer Categories',
            'icon': 'fas fa-gift',
            'badge': 'Combo Offer',
            'add_label': 'Add Combo Category',
        },
        'accessory': {
            'title': 'Accessory Categories',
            'icon': 'fas fa-toolbox',
            'badge': 'Accessory',
            'add_label': 'Add Accessory Category',
        },
        'plant': {
            'title': 'Plant Categories',
            'icon': 'fas fa-seedling',
            'badge': 'Plant',
            'add_label': 'Add Plant Category',
        },
    }

    all_categories_qs = Category.objects.all()
    counts_by_type = {key: all_categories_qs.filter(category_type=key).count() for key in type_configs.keys()}

    search_query = request.GET.get('search', '').strip()
    type_filter = request.GET.get('category_type', '').strip()

    categories_qs = all_categories_qs
    if type_filter and type_filter in type_configs:
        categories_qs = categories_qs.filter(category_type=type_filter)
    else:
        type_filter = ''

    if search_query:
        categories_qs = categories_qs.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )

    categories = list(categories_qs.order_by('name'))

    filtered_counts_by_type = {key: categories_qs.filter(category_type=key).count() for key in type_configs.keys()}

    category_cards = []
    for category in categories:
        config = type_configs.get(category.category_type, {})
        category_cards.append({
            'object': category,
            'icon': config.get('icon', 'fas fa-tags'),
            'badge': config.get('badge', category.get_category_type_display()),
        })

    type_actions = []
    for type_key, config in type_configs.items():
        type_actions.append({
            'key': type_key,
            'title': config['title'],
            'icon': config['icon'],
            'count': filtered_counts_by_type.get(type_key, 0),
            'badge': config.get('badge', config['title']),
            'total_count': counts_by_type.get(type_key, 0),
            'is_selected': type_filter == type_key,
        })

    total_categories = len(categories)
    overall_category_count = all_categories_qs.count()
    has_any_categories = overall_category_count > 0
    add_category_url = reverse('add_category')
    add_category_label = 'Add Category'
    if type_filter:
        add_category_label = type_configs[type_filter]['add_label']
        add_category_url = f"{add_category_url}?type={type_filter}"

    context = {
        'category_cards': category_cards,
        'type_actions': type_actions,
        'has_any_categories': has_any_categories,
        'total_categories': total_categories,
        'overall_category_count': overall_category_count,
        'type_config_map': type_configs,
        'add_category_url': add_category_url,
        'add_category_label': add_category_label,
        'search_query': search_query,
        'selected_type': type_filter,
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        summary_html = render_to_string(
            'store/staff/partials/category_summary.html',
            context,
            request=request,
        )
        cards_html = render_to_string(
            'store/staff/partials/category_cards.html',
            context,
            request=request,
        )
        return JsonResponse({
            'summary_html': summary_html,
            'cards_html': cards_html,
        })

    return render(request, 'store/staff/categories.html', context)


@login_required
@user_passes_test(is_staff)
def staff_plants_view(request):
    categories = Category.objects.filter(category_type='plant').order_by('name')
    plants_qs = Plant.objects.select_related('category').order_by('display_order', 'name')

    category_id = request.GET.get('category', '').strip()
    search_query = request.GET.get('search', '').strip()
    availability = request.GET.get('availability', '').strip()

    if category_id and category_id.isdigit():
        plants_qs = plants_qs.filter(category_id=int(category_id))
    else:
        category_id = ''

    if search_query:
        plants_qs = plants_qs.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )

    if availability == 'active':
        plants_qs = plants_qs.filter(is_active=True)
    elif availability == 'inactive':
        plants_qs = plants_qs.filter(is_active=False)
    elif availability == 'in_stock':
        plants_qs = plants_qs.filter(stock_quantity__gt=0)
    elif availability == 'out_of_stock':
        plants_qs = plants_qs.filter(stock_quantity__lte=0)
    else:
        availability = ''

    plants = list(plants_qs)

    context = {
        'plants': plants,
        'categories': categories,
        'selected_category_id': category_id,
        'search_query': search_query,
        'availability': availability,
        'total_plants': len(plants),
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render_to_string('store/staff/_plants_results.html', context, request=request)
        return JsonResponse({'html': html, 'total': len(plants)})

    return render(request, 'store/staff/plants.html', context)


@login_required
@user_passes_test(is_staff)
def add_plant_view(request):
    if request.method == 'POST':
        form = PlantForm(request.POST, request.FILES)
        if form.is_valid():
            plant = form.save(commit=False)
            if not plant.created_by:
                plant.created_by = request.user
            plant.save()
            messages.success(request, 'Plant added successfully.')
            return redirect('staff_plants')
    else:
        initial = {}
        category_hint = request.GET.get('category')
        if category_hint and category_hint.isdigit():
            try:
                plant_category = Category.objects.get(id=int(category_hint), category_type='plant')
                initial['category'] = plant_category.id
            except Category.DoesNotExist:
                pass
        form = PlantForm(initial=initial)

    return render(request, 'store/staff/add_plant.html', {'form': form})


@login_required
@user_passes_test(is_staff)
def edit_plant_view(request, plant_id):
    plant = get_object_or_404(Plant, id=plant_id)
    if request.method == 'POST':
        form = PlantForm(request.POST, request.FILES, instance=plant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Plant updated successfully.')
            return redirect('staff_plants')
    else:
        form = PlantForm(instance=plant)

    return render(request, 'store/staff/add_plant.html', {'form': form, 'plant': plant})


@login_required
@user_passes_test(is_staff)
@require_POST
def delete_plant_view(request, plant_id):
    plant = get_object_or_404(Plant, id=plant_id)
    plant.delete()
    messages.success(request, 'Plant deleted successfully.')
    return redirect('staff_plants')


@login_required
@user_passes_test(is_staff)
def delete_category_view(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    category.delete()
    messages.success(request, 'Category deleted successfully.')
    return redirect('staff_categories')


@login_required
@user_passes_test(is_staff)
def add_breed_view(request):
    if request.method == 'POST':
        form = BreedForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Breed added successfully.')
            return redirect('staff_breeds')
    else:
        form = BreedForm()
    
    categories = Category.objects.all()
    return render(request, 'store/staff/add_breed.html', {'form': form, 'categories': categories})


@login_required
@user_passes_test(is_staff)
def staff_breeds_view(request):
    breeds = Breed.objects.all()
    return render(request, 'store/staff/breeds.html', {'breeds': breeds})


@login_required
@user_passes_test(is_staff)
def delete_breed_view(request, breed_id):
    breed = get_object_or_404(Breed, id=breed_id)
    breed.delete()
    messages.success(request, 'Breed deleted successfully.')
    return redirect('staff_breeds')


@login_required
@user_passes_test(is_staff)
def add_fish_view(request):
    if request.method == 'POST':
        # Check if this is a breed creation request
        if 'create_breed' in request.POST:
            breed_form = BreedForm(request.POST)
            if breed_form.is_valid():
                new_breed = breed_form.save()
                # Return JSON response for AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'breed_id': new_breed.id,
                        'breed_name': new_breed.name,
                        'message': 'Breed created successfully!'
                    })
                messages.success(request, 'Breed created successfully.')
        
        # Normal fish creation
        form = FishForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fish added successfully.')
            return redirect('staff_fish_list')
    else:
        form = FishForm()
    
    breed_form = BreedForm()
    categories = Category.objects.all()
    breeds = Breed.objects.all()
    return render(request, 'store/staff/add_fish.html', {
        'form': form,
        'breed_form': breed_form,
        'categories': categories,
        'breeds': breeds,
    })


@login_required
@user_passes_test(is_staff)
def edit_fish_view(request, fish_id):
    fish = get_object_or_404(Fish, id=fish_id)
    
    if request.method == 'POST':
        # Check if this is a breed creation request
        if 'create_breed' in request.POST:
            breed_form = BreedForm(request.POST)
            if breed_form.is_valid():
                new_breed = breed_form.save()
                # Return JSON response for AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'breed_id': new_breed.id,
                        'breed_name': new_breed.name,
                        'message': 'Breed created successfully!'
                    })
                messages.success(request, 'Breed created successfully.')
        
        # Normal fish update
        form = FishForm(request.POST, request.FILES, instance=fish)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fish updated successfully.')
            return redirect('staff_fish_list')
    else:
        form = FishForm(instance=fish)
    
    breed_form = BreedForm()
    categories = Category.objects.all()
    breeds = Breed.objects.all()
    return render(request, 'store/staff/edit_fish.html', {
        'form': form,
        'breed_form': breed_form,
        'fish': fish,
        'categories': categories,
        'breeds': breeds,
    })


@login_required
@user_passes_test(is_staff)
def delete_fish_view(request, fish_id):
    fish = get_object_or_404(Fish, id=fish_id)
    fish.delete()
    messages.success(request, 'Fish deleted successfully.')
    return redirect('staff_fish_list')


# Staff: Manage Fish Media
@login_required
@user_passes_test(is_staff)
def staff_fish_media_view(request, fish_id):
    fish = get_object_or_404(Fish, id=fish_id)
    media_items = FishMedia.objects.filter(fish=fish)

    if request.method == 'POST':
        form = FishMediaForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.fish = fish
            # Require either file or external URL
            if not obj.file and not obj.external_url:
                messages.error(request, 'Please provide a file or an external URL.')
            else:
                obj.save()
                messages.success(request, 'Media added successfully.')
                return redirect('staff_fish_media', fish_id=fish.id)
    else:
        form = FishMediaForm()

    return render(request, 'store/staff/fish_media.html', {
        'fish': fish,
        'media_items': media_items,
        'form': form,
    })


@login_required
@user_passes_test(is_staff)
def staff_delete_fish_media_view(request, media_id):
    media = get_object_or_404(FishMedia, id=media_id)
    fish_id = media.fish.id
    media.delete()
    messages.success(request, 'Media deleted successfully.')
    return redirect('staff_fish_media', fish_id=fish_id)


@login_required
@user_passes_test(is_staff)
def staff_edit_fish_media_view(request, media_id):
    media = get_object_or_404(FishMedia, id=media_id)
    fish = media.fish

    if request.method == 'POST':
        form = FishMediaForm(request.POST, request.FILES, instance=media)
        if form.is_valid():
            obj = form.save(commit=False)
            if not obj.file and not obj.external_url:
                messages.error(request, 'Please provide a file or an external URL.')
            else:
                obj.save()
                messages.success(request, 'Media updated successfully.')
                return redirect('staff_fish_media', fish_id=fish.id)
    else:
        form = FishMediaForm(instance=media)

    return render(request, 'store/staff/edit_fish_media.html', {
        'fish': fish,
        'media': media,
        'form': form,
    })


# Admin Views
@login_required
@user_passes_test(is_admin)
def admin_dashboard_view(request):
    total_users = CustomUser.objects.filter(role='customer').count()
    total_staff = CustomUser.objects.filter(role='staff').count()
    total_orders = Order.objects.count()
    total_revenue = Order.objects.filter(status='delivered').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    recent_orders = Order.objects.all().order_by('-created_at')[:5]
    
    # Get data for charts
    # Last 6 months of data
    from django.utils import timezone
    six_months_ago = timezone.now() - timedelta(days=180)
    
    # Monthly orders data
    monthly_orders = (Order.objects
        .filter(created_at__gte=six_months_ago)
        .values('created_at__year', 'created_at__month')
        .annotate(count=Count('id'))
        .order_by('created_at__year', 'created_at__month'))
    
    # Monthly revenue data
    monthly_revenue = (Order.objects
        .filter(created_at__gte=six_months_ago, status='delivered')
        .values('created_at__year', 'created_at__month')
        .annotate(total=Sum('total_amount'))
        .order_by('created_at__year', 'created_at__month'))
    
    # Convert month strings to datetime for proper formatting
    months = []
    orders_data = []
    revenue_data = []
    
    # Create a mapping of previous 6 months
    month_data = {}
    current = timezone.now()
    for i in range(6):
        month_key = (current.year, current.month)
        month_data[month_key] = {
            'month_name': current.strftime('%B %Y'),
            'orders': 0,
            'revenue': 0
        }
        # Go to previous month
        if current.month == 1:
            current = current.replace(year=current.year-1, month=12, day=1)
        else:
            current = current.replace(month=current.month-1, day=1)
    
    # Fill in actual data
    for entry in monthly_orders:
        month_key = (entry['created_at__year'], entry['created_at__month'])
        if month_key in month_data:
            month_data[month_key]['orders'] = entry['count']
    
    for entry in monthly_revenue:
        month_key = (entry['created_at__year'], entry['created_at__month'])
        if month_key in month_data:
            month_data[month_key]['revenue'] = float(entry['total'] if entry['total'] else 0)
    
    # Convert to lists for the template
    sorted_months = sorted(month_data.keys())  # Oldest first for proper chart display
    for month_key in sorted_months:
        months.append(month_data[month_key]['month_name'])
        orders_data.append(month_data[month_key]['orders'])
        revenue_data.append(month_data[month_key]['revenue'])
    
    return render(request, 'store/admin/dashboard.html', {
        'total_users': total_users,
        'total_staff': total_staff,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'recent_orders': recent_orders,
        'chart_months': months,
        'chart_orders': orders_data,
        'chart_revenue': revenue_data,
    })


@login_required
@user_passes_test(is_admin)
def admin_staff_list_view(request):
    staff_members = CustomUser.objects.filter(role='staff')
    return render(request, 'store/admin/staff_list.html', {'staff_members': staff_members})


@login_required
@user_passes_test(is_admin)
def add_staff_view(request):
    if request.method == 'POST':
        # Staff creation with admin-provided password (password fields are on the form)
        form = StaffCreateForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'staff'
            user.email_verified = True
            user.is_active = True
            # Set the password provided by the admin
            password = form.cleaned_data.get('password1')
            if password:
                user.set_password(password)
            else:
                user.set_unusable_password()
            user.save()
            # Send a simple welcome email (do not include passwords in email)
            try:
                send_mail(
                    f'You have been added as Staff - {settings.SITE_NAME}',
                    (
                        f"Hello {user.username},\n\n"
                        f"You've been added as staff to {settings.SITE_NAME}.\n\n"
                        "You can now login using your account credentials. If you did not set a password, please use the password reset flow.\n\n"
                        f"Thanks,\n{settings.SITE_NAME}"
                    ),
                    settings.DEFAULT_FROM_EMAIL or 'noreply@aquafishstore.com',
                    [user.email],
                    fail_silently=False,
                )
                messages.success(request, 'Staff member added and notification email sent successfully.')
            except Exception as e:
                messages.warning(request, f'Staff member added, but failed to send email: {e}')
            return redirect('admin_staff_list')
    else:
        form = StaffCreateForm()
    
    return render(request, 'store/admin/add_staff.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def remove_staff_view(request, user_id):
    staff = get_object_or_404(CustomUser, id=user_id, role='staff')
    staff.delete()
    messages.success(request, 'Staff member removed successfully.')
    return redirect('admin_staff_list')


@login_required
@user_passes_test(is_admin)
def block_staff_view(request, user_id):
    staff = get_object_or_404(CustomUser, id=user_id, role='staff')
    staff.is_blocked = True
    staff.save()
    # Terminate any active sessions for this staff member so they are immediately logged out
    removed = terminate_user_sessions(staff)
    if removed:
        messages.success(request, f'Staff member {staff.username} has been blocked and {removed} active session(s) terminated.')
    else:
        messages.success(request, f'Staff member {staff.username} has been blocked.')
    return redirect('admin_staff_list')


@login_required
@user_passes_test(is_admin)
def unblock_staff_view(request, user_id):
    staff = get_object_or_404(CustomUser, id=user_id, role='staff')
    staff.is_blocked = False
    staff.save()
    messages.success(request, f'Staff member {staff.username} has been unblocked.')
    return redirect('admin_staff_list')


@login_required
@user_passes_test(is_admin)
def admin_categories_view(request):
    type_configs = {
        'fish': {
            'title': 'Fish Categories',
            'icon': 'fas fa-fish',
            'add_label': 'Add Fish Category',
            'badge': 'Fish',
        },
        'combo': {
            'title': 'Combo Offer Categories',
            'icon': 'fas fa-gift',
            'add_label': 'Add Combo Offer Category',
            'badge': 'Combo Offer',
        },
        'accessory': {
            'title': 'Accessory Categories',
            'icon': 'fas fa-toolbox',
            'add_label': 'Add Accessory Category',
            'badge': 'Accessory',
        },
        'plant': {
            'title': 'Plant Categories',
            'icon': 'fas fa-seedling',
            'add_label': 'Add Plant Category',
            'badge': 'Plant',
        },
    }

    # Sanitize function for cleaning UTF-8 errors
    def sanitize_string(value):
        """Remove invalid UTF-8 characters from string."""
        if not value:
            return value
        try:
            # Try to encode and decode to ensure valid UTF-8
            if isinstance(value, bytes):
                return value.decode('utf-8', errors='ignore')
            return value.encode('utf-8', errors='ignore').decode('utf-8')
        except Exception:
            return str(value).encode('utf-8', errors='ignore').decode('utf-8', errors='ignore')

    all_categories_qs = Category.objects.all()
    counts_by_type = {key: all_categories_qs.filter(category_type=key).count() for key in type_configs.keys()}

    search_query = request.GET.get('search', '').strip()
    type_filter = request.GET.get('category_type', '').strip()

    categories_qs = all_categories_qs
    if type_filter and type_filter in type_configs:
        categories_qs = categories_qs.filter(category_type=type_filter)

    if search_query:
        categories_qs = categories_qs.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )

    # Fetch and sanitize all categories
    categories = []
    for category in categories_qs.order_by('name'):
        try:
            # Sanitize category fields before using them
            category.name = sanitize_string(category.name)
            category.description = sanitize_string(category.description)
            categories.append(category)
        except Exception:
            # Skip categories that can't be sanitized
            continue

    filtered_counts_by_type = {key: categories_qs.filter(category_type=key).count() for key in type_configs.keys()}

    category_cards = []
    for category in categories:
        config = type_configs.get(category.category_type, {})
        category_cards.append({
            'object': category,
            'icon': config.get('icon', 'fas fa-tags'),
            'badge': config.get('badge', category.get_category_type_display()),
            'title': config.get('title', category.get_category_type_display()),
        })

    type_actions = []
    for type_key, config in type_configs.items():
        type_actions.append({
            'key': type_key,
            'title': config['title'],
            'icon': config['icon'],
            'add_label': config['add_label'],
            'add_url': f"{reverse('admin_add_category')}?type={type_key}",
            'count': filtered_counts_by_type.get(type_key, 0),
            'badge': config.get('badge', config['title']),
            'total_count': counts_by_type.get(type_key, 0),
            'is_selected': type_filter == type_key,
        })

    total_categories = len(categories)
    overall_category_count = all_categories_qs.count()
    has_any_categories = overall_category_count > 0
    add_category_url = reverse('admin_add_category')
    add_category_label = 'Add Category'
    if type_filter in type_configs:
        add_category_label = type_configs[type_filter]['add_label']
        add_category_url = f"{add_category_url}?type={type_filter}"

    context = {
        'category_cards': category_cards,
        'type_actions': type_actions,
        'has_any_categories': has_any_categories,
        'total_categories': total_categories,
        'overall_category_count': overall_category_count,
        'type_config_map': type_configs,
        'add_category_url': add_category_url,
        'add_category_label': add_category_label,
        'search_query': sanitize_string(search_query),
        'selected_type': sanitize_string(type_filter if type_filter in type_configs else ''),
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        summary_html = render_to_string(
            'store/admin/partials/category_summary.html',
            context,
            request=request,
        )
        cards_html = render_to_string(
            'store/admin/partials/category_cards.html',
            context,
            request=request,
        )
        return JsonResponse({
            'summary_html': summary_html,
            'cards_html': cards_html,
        })

    return render(
        request,
        'store/admin/categories.html',
        context,
    )


@login_required
@user_passes_test(is_admin)
def admin_plants_view(request):
    categories = Category.objects.filter(category_type='plant').order_by('name')
    plants_qs = Plant.objects.select_related('category').order_by('display_order', 'name')

    category_id = request.GET.get('category', '').strip()
    search_query = request.GET.get('search', '').strip()
    availability = request.GET.get('availability', '').strip()

    if category_id and category_id.isdigit():
        plants_qs = plants_qs.filter(category_id=int(category_id))
    else:
        category_id = ''

    if search_query:
        plants_qs = plants_qs.filter(Q(name__icontains=search_query) | Q(description__icontains=search_query))

    if availability == 'active':
        plants_qs = plants_qs.filter(is_active=True)
    elif availability == 'inactive':
        plants_qs = plants_qs.filter(is_active=False)
    elif availability == 'in_stock':
        plants_qs = plants_qs.filter(stock_quantity__gt=0)
    elif availability == 'out_of_stock':
        plants_qs = plants_qs.filter(stock_quantity__lte=0)
    else:
        availability = ''

    plants = list(plants_qs)

    context = {
        'plants': plants,
        'categories': categories,
        'selected_category_id': category_id,
        'search_query': search_query,
        'availability': availability,
        'total_plants': len(plants),
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render_to_string('store/admin/_plants_results.html', context, request=request)
        return JsonResponse({'html': html, 'total': len(plants)})

    return render(request, 'store/admin/plants.html', context)


@login_required
@user_passes_test(is_admin)
def admin_add_plant_view(request):
    if request.method == 'POST':
        form = PlantForm(request.POST, request.FILES)
        if form.is_valid():
            plant = form.save(commit=False)
            if not plant.created_by:
                plant.created_by = request.user
            plant.save()
            messages.success(request, 'Plant created successfully.')
            return redirect('admin_plants')
    else:
        initial = {}
        category_hint = request.GET.get('category')
        if category_hint and category_hint.isdigit():
            try:
                plant_category = Category.objects.get(id=int(category_hint), category_type='plant')
                initial['category'] = plant_category.id
            except Category.DoesNotExist:
                pass
        form = PlantForm(initial=initial)

    return render(request, 'store/admin/add_plant.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def admin_edit_plant_view(request, plant_id):
    plant = get_object_or_404(Plant, id=plant_id)
    if request.method == 'POST':
        form = PlantForm(request.POST, request.FILES, instance=plant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Plant updated successfully.')
            return redirect('admin_plants')
    else:
        form = PlantForm(instance=plant)

    return render(request, 'store/admin/add_plant.html', {'form': form, 'plant': plant})


@login_required
@user_passes_test(is_admin)
@require_POST
def admin_delete_plant_view(request, plant_id):
    plant = get_object_or_404(Plant, id=plant_id)
    plant.delete()
    messages.success(request, 'Plant deleted successfully.')
    return redirect('admin_plants')


@login_required
@user_passes_test(is_admin)
def admin_add_category_view(request):
    type_hint = request.GET.get('type') or request.POST.get('category_type')
    valid_types = dict(Category.CATEGORY_TYPES)
    type_label = valid_types.get(type_hint) if type_hint in valid_types else None

    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category added successfully.')
            return redirect('admin_categories')
    else:
        initial = {}
        if type_hint in valid_types:
            initial['category_type'] = type_hint
        form = CategoryForm(initial=initial)

    add_heading = 'Add Category'
    if type_label:
        add_heading = f"Add {type_label} Category"

    return render(
        request,
        'store/admin/add_category.html',
        {
            'form': form,
            'add_heading': add_heading,
            'type_hint': type_hint if type_label else None,
            'type_label': type_label,
        },
    )


@login_required
@user_passes_test(is_admin)
def admin_edit_category_view(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated successfully.')
            return redirect('admin_categories')
    else:
        form = CategoryForm(instance=category)
    
    return render(
        request,
        'store/admin/edit_category.html',
        {
            'form': form,
            'category': category,
        },
    )


@login_required
@user_passes_test(is_admin)
def admin_delete_category_view(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    category.delete()
    messages.success(request, 'Category deleted successfully.')
    return redirect('admin_categories')


@login_required
@user_passes_test(is_admin)
def admin_breeds_view(request):
    breeds = Breed.objects.all()
    return render(request, 'store/admin/breeds.html', {'breeds': breeds})


@login_required
@user_passes_test(is_admin)
def admin_add_breed_view(request):
    if request.method == 'POST':
        form = BreedForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Breed added successfully.')
            return redirect('admin_breeds')
    else:
        form = BreedForm()
    
    categories = Category.objects.all()
    return render(request, 'store/admin/add_breed.html', {'form': form, 'categories': categories})


@login_required
@user_passes_test(is_admin)
def admin_delete_breed_view(request, breed_id):
    breed = get_object_or_404(Breed, id=breed_id)
    breed.delete()
    messages.success(request, 'Breed deleted successfully.')
    return redirect('admin_breeds')


@login_required
@user_passes_test(is_admin)
def admin_fishes_view(request):
    fishes = Fish.objects.select_related('category', 'breed').all().order_by('-is_featured', '-created_at')

    search_query = (request.GET.get('search') or '').strip()
    category_id = (request.GET.get('category') or '').strip()
    breed_id = (request.GET.get('breed') or '').strip()
    availability = (request.GET.get('availability') or '').strip()
    featured = (request.GET.get('featured') or '').strip()

    if search_query:
        fishes = fishes.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if category_id:
        fishes = fishes.filter(category_id=category_id)

    if breed_id:
        fishes = fishes.filter(breed_id=breed_id)

    if availability == 'available':
        fishes = fishes.filter(is_available=True)
    elif availability == 'unavailable':
        fishes = fishes.filter(is_available=False)
    elif availability == 'in_stock':
        fishes = fishes.filter(stock_quantity__gt=0)
    elif availability == 'out_of_stock':
        fishes = fishes.filter(stock_quantity__lte=0)

    if featured == 'yes':
        fishes = fishes.filter(is_featured=True)
    elif featured == 'no':
        fishes = fishes.filter(is_featured=False)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render_to_string('store/admin/_fishes_results.html', {'fishes': fishes}, request=request)
        return JsonResponse({'html': html})

    categories = Category.objects.filter(category_type='fish').order_by('name')
    return render(request, 'store/admin/fishes.html', {
        'fishes': fishes,
        'categories': categories,
        'search_query': search_query,
        'category_id': category_id,
        'availability': availability,
        'featured': featured,
    })


@login_required
@user_passes_test(is_admin)
def admin_add_fish_view(request):
    if request.method == 'POST':
        # Check if this is a breed creation request
        if 'create_breed' in request.POST:
            breed_form = BreedForm(request.POST)
            if breed_form.is_valid():
                new_breed = breed_form.save()
                # Return JSON response for AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'breed_id': new_breed.id,
                        'breed_name': new_breed.name,
                        'message': 'Breed created successfully!'
                    })
                messages.success(request, 'Breed created successfully.')
        
        # Normal fish creation
        form = FishForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fish added successfully.')
            return redirect('admin_fishes')
    else:
        form = FishForm()
    
    breed_form = BreedForm()
    categories = Category.objects.all()
    breeds = Breed.objects.all()
    return render(request, 'store/admin/add_fish.html', {
        'form': form,
        'breed_form': breed_form,
        'categories': categories,
        'breeds': breeds,
    })


@login_required
@user_passes_test(is_admin)
def admin_delete_fish_view(request, fish_id):
    fish = get_object_or_404(Fish, id=fish_id)
    fish.delete()
    messages.success(request, 'Fish deleted successfully.')
    return redirect('admin_fishes')


# -------------------- Services (Admin Managed) --------------------
@login_required
@user_passes_test(is_admin)
def admin_services_view(request):
    services = Service.objects.all()
    return render(request, 'store/admin/services.html', {'services': services})


@login_required
@user_passes_test(is_admin)
def admin_add_service_view(request):
    if request.method == 'POST':
        form = ServiceForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Service added successfully.')
            return redirect('admin_services')
    else:
        form = ServiceForm()
    return render(request, 'store/admin/add_service.html', {'form': form})


@login_required
@user_passes_test(is_staff)
def staff_accessories_view(request):
    from .models import Accessory
    # Base queryset with related category for efficiency
    accessories = Accessory.objects.select_related('category').all()

    # Parse filters
    search_query = (request.GET.get('search') or '').strip()
    category_id = (request.GET.get('category') or '').strip()
    availability = (request.GET.get('availability') or '').strip()

    # Apply search filter
    if search_query:
        accessories = accessories.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Apply category filter
    if category_id:
        accessories = accessories.filter(category_id=category_id)

    # Apply availability/status filters (reuse fish values for parity)
    if availability == 'available':
        # Treat as active
        accessories = accessories.filter(is_active=True)
    elif availability == 'unavailable':
        accessories = accessories.filter(is_active=False)
    elif availability == 'in_stock':
        accessories = accessories.filter(stock_quantity__gt=0)
    elif availability == 'out_of_stock':
        accessories = accessories.filter(stock_quantity__lte=0)

    # AJAX: return only table HTML for partial update
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render_to_string('store/staff/_accessories_table.html', {
            'accessories': accessories,
        }, request=request)
        if not accessories:
            empty_html = render_to_string('store/staff/_accessories_empty.html', request=request)
            return JsonResponse({'html': html + empty_html})
        return JsonResponse({'html': html})

    # Categories with counts for filter dropdown
    categories_with_count = Category.objects.filter(category_type='accessory').annotate(
        accessory_count=Count('accessories', distinct=True)
    ).order_by('name')

    return render(request, 'store/staff/accessories.html', {
        'accessories': accessories,
        'search_query': search_query,
        'categories': categories_with_count,
        'category_id': category_id,
        'availability': availability,
    })


@login_required
@user_passes_test(is_staff)
def add_accessory_view(request):
    from .models import Accessory
    from .forms import AccessoryForm

    if request.method == 'POST':
        form = AccessoryForm(request.POST, request.FILES)
        if form.is_valid():
            accessory = form.save(commit=False)
            accessory.created_by = request.user
            accessory.save()
            messages.success(request, 'Accessory added successfully.')
            return redirect('staff_accessories')
    else:
        form = AccessoryForm()
    return render(request, 'store/staff/add_accessory.html', {'form': form})


@login_required
@user_passes_test(is_staff)
def accessories_add_view(request):
    """Unified add-accessory page available to staff and admin (uses staff template)."""
    from .forms import AccessoryForm

    if request.method == 'POST':
        form = AccessoryForm(request.POST, request.FILES)
        if form.is_valid():
            accessory = form.save(commit=False)
            accessory.created_by = request.user
            accessory.save()
            messages.success(request, 'Accessory added successfully.')
            # Role-aware redirect
            if hasattr(request.user, 'role') and request.user.role == 'admin':
                return redirect('admin_accessories')
            return redirect('staff_accessories')
    else:
        form = AccessoryForm()

    # Reuse staff add template for simplicity
    return render(request, 'store/staff/add_accessory.html', {'form': form})


@login_required
@user_passes_test(is_staff)
def edit_accessory_view(request, accessory_id):
    from .models import Accessory
    from .forms import AccessoryForm
    accessory = get_object_or_404(Accessory, id=accessory_id)
    if request.method == 'POST':
        form = AccessoryForm(request.POST, request.FILES, instance=accessory)
        if form.is_valid():
            form.save()
            messages.success(request, 'Accessory updated successfully.')
            return redirect('staff_accessories')
    else:
        form = AccessoryForm(instance=accessory)
    return render(request, 'store/staff/edit_accessory.html', {'form': form, 'accessory': accessory})


@login_required
@user_passes_test(is_staff)
@require_POST
def delete_accessory_view(request, accessory_id):
    from .models import Accessory
    accessory = get_object_or_404(Accessory, id=accessory_id)
    accessory.delete()
    messages.success(request, 'Accessory deleted successfully.')
    return redirect('staff_accessories')


def _build_admin_accessories_context(request):
    from .models import Accessory

    search_query = (request.GET.get('search') or '').strip()
    category_id = (request.GET.get('category') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()

    qs = Accessory.objects.all().select_related('category')
    total_count = qs.count()

    if search_query:
        qs = qs.filter(Q(name__icontains=search_query) | Q(description__icontains=search_query))

    if category_id:
        qs = qs.filter(category_id=category_id)

    if status_filter == 'active':
        qs = qs.filter(is_active=True)
    elif status_filter == 'inactive':
        qs = qs.filter(is_active=False)
    elif status_filter == 'in_stock':
        qs = qs.filter(stock_quantity__gt=0)
    elif status_filter == 'out_of_stock':
        qs = qs.filter(stock_quantity__lte=0)

    qs = qs.order_by('display_order', 'name')
    filtered_count = qs.count()

    categories = Category.objects.filter(category_type='accessory').order_by('name')
    filters_active = bool(search_query or category_id or status_filter)

    return {
        'accessories': qs,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_status': status_filter,
        'categories': categories,
        'filtered_count': filtered_count,
        'total_count': total_count,
        'filters_active': filters_active,
    }


@login_required
@user_passes_test(is_admin)
def admin_accessories_view(request):
    context = _build_admin_accessories_context(request)
    return render(request, 'store/admin/accessories.html', context)


@login_required
@user_passes_test(is_admin)
@require_GET
def admin_accessories_ajax_view(request):
    context = _build_admin_accessories_context(request)
    table_html = render_to_string('store/admin/_accessories_table.html', context, request=request)
    return JsonResponse({
        'table_html': table_html,
        'filtered_count': context.get('filtered_count', 0),
        'total_count': context.get('total_count', 0),
        'filters_active': context.get('filters_active', False),
    })


@login_required
@user_passes_test(is_admin)
def admin_add_accessory_view(request):
    from .forms import AccessoryForm
    if request.method == 'POST':
        form = AccessoryForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, 'Accessory created successfully.')
            return redirect('admin_accessories')
    else:
        form = AccessoryForm()
    return render(request, 'store/admin/add_accessory.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def admin_edit_accessory_view(request, accessory_id):
    from .models import Accessory
    from .forms import AccessoryForm
    accessory = get_object_or_404(Accessory, id=accessory_id)
    if request.method == 'POST':
        form = AccessoryForm(request.POST, request.FILES, instance=accessory)
        if form.is_valid():
            form.save()
            messages.success(request, 'Accessory updated successfully.')
            return redirect('admin_accessories')
    else:
        form = AccessoryForm(instance=accessory)
    return render(request, 'store/admin/add_accessory.html', {'form': form, 'accessory': accessory})


@login_required
@user_passes_test(is_admin)
@require_POST
def admin_delete_accessory_view(request, accessory_id):
    from .models import Accessory
    accessory = get_object_or_404(Accessory, id=accessory_id)
    accessory.delete()
    messages.success(request, 'Accessory deleted successfully.')
    return redirect('admin_accessories')


@login_required
@user_passes_test(is_admin)
def admin_edit_service_view(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == 'POST':
        form = ServiceForm(request.POST, request.FILES, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, 'Service updated successfully.')
            return redirect('admin_services')
    else:
        form = ServiceForm(instance=service)
    return render(request, 'store/admin/add_service.html', {'form': form, 'is_edit': True})


@login_required
@user_passes_test(is_admin)
def admin_delete_service_view(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    service.delete()
    messages.success(request, 'Service deleted successfully.')
    return redirect('admin_services')


@login_required
@user_passes_test(is_admin)
def admin_blogs_view(request):
    from .models import BlogPost
    posts = BlogPost.objects.all().order_by('-created_at')
    return render(request, 'store/admin/blogs.html', {'posts': posts})


@login_required
@user_passes_test(is_admin)
def admin_add_blog_view(request):
    from .forms import BlogPostForm
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.author = request.user
            # If published flag set and no published_at, set now
            if obj.is_published and not obj.published_at:
                obj.published_at = timezone.now()
            obj.save()
            messages.success(request, 'Blog post created successfully.')
            return redirect('admin_blogs')
    else:
        form = BlogPostForm()
    return render(request, 'store/admin/add_blog.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def admin_edit_blog_view(request, post_id):
    from .models import BlogPost
    from .forms import BlogPostForm
    post = get_object_or_404(BlogPost, id=post_id)
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            obj = form.save(commit=False)
            if obj.is_published and not obj.published_at:
                obj.published_at = timezone.now()
            obj.save()
            messages.success(request, 'Blog post updated successfully.')
            return redirect('admin_blogs')
    else:
        form = BlogPostForm(instance=post)
    return render(request, 'store/admin/add_blog.html', {'form': form, 'post': post})


@login_required
@user_passes_test(is_admin)
def admin_delete_blog_view(request, post_id):
    from .models import BlogPost
    post = get_object_or_404(BlogPost, id=post_id)
    post.delete()
    messages.success(request, 'Blog post deleted.')
    return redirect('admin_blogs')


# -------------------- Contact Info (Admin Managed) --------------------
@login_required
@user_passes_test(is_admin)
def admin_shipping_charges_view(request):
    setting, _ = ShippingChargeSetting.objects.get_or_create(
        key='default',
        defaults={'kerala_rate': Decimal('60.00'), 'default_rate': Decimal('100.00')},
    )
    locations = ShippingChargeByLocation.objects.order_by('location_name')
    form = ShippingChargeForm(instance=setting)
    location_form = ShippingChargeByLocationForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_location':
            location_form = ShippingChargeByLocationForm(request.POST)
            if location_form.is_valid():
                location_form.save()
                _get_shipping_rates.cache_clear()
                messages.success(request, 'Location-based shipping charge added.')
                return redirect('admin_shipping_charges')
        elif action == 'edit_location':
            location_id = request.POST.get('location_id')
            if location_id:
                location = ShippingChargeByLocation.objects.filter(id=location_id).first()
                if location:
                    edit_form = ShippingChargeByLocationForm(request.POST, instance=location)
                    if edit_form.is_valid():
                        edit_form.save()
                        _get_shipping_rates.cache_clear()
                        messages.success(request, 'Location-based shipping charge updated.')
                        return redirect('admin_shipping_charges')
        elif action == 'delete_location':
            location_id = request.POST.get('location_id')
            if location_id:
                ShippingChargeByLocation.objects.filter(id=location_id).delete()
                _get_shipping_rates.cache_clear()
                messages.success(request, 'Location-based shipping charge removed.')
                return redirect('admin_shipping_charges')
        else:
            form = ShippingChargeForm(request.POST, instance=setting)
            if form.is_valid():
                form.save()
                _get_shipping_rates.cache_clear()
                messages.success(request, 'Shipping charges updated successfully.')
                return redirect('admin_shipping_charges')
    else:
        form = ShippingChargeForm(instance=setting)

    context = {
        'form': form,
        'setting': setting,
        'last_updated': setting.updated_at,
        'location_form': location_form,
        'locations': locations,
    }
    return render(request, 'store/admin/shipping_charges.html', context)


@login_required
@user_passes_test(is_admin)
def admin_contact_view(request):
    contact = ContactInfo.objects.first()
    return render(request, 'store/admin/contact.html', {'contact': contact})


@login_required
@user_passes_test(is_admin)
def admin_add_contact_view(request):
    existing = ContactInfo.objects.first()
    if existing:
        return redirect('admin_edit_contact', contact_id=existing.id)
    if request.method == 'POST':
        form = ContactInfoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contact information saved.')
            return redirect('admin_contact')
    else:
        form = ContactInfoForm()
    return render(request, 'store/admin/add_contact.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def admin_edit_contact_view(request, contact_id):
    contact = get_object_or_404(ContactInfo, id=contact_id)
    if request.method == 'POST':
        form = ContactInfoForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contact information updated.')
            return redirect('admin_contact')
    else:
        form = ContactInfoForm(instance=contact)
    return render(request, 'store/admin/add_contact.html', {'form': form, 'is_edit': True})


@login_required
@user_passes_test(is_admin)
def admin_gallery_view(request):
    contact = ContactInfo.objects.first()
    if not contact:
        messages.error(request, 'Please create Contact Information before adding gallery items.')
        return redirect('admin_add_contact')

    from .forms import ContactGalleryForm
    media_items = ContactGalleryMedia.objects.filter(contact=contact)

    if request.method == 'POST':
        form = ContactGalleryForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.contact = contact
            if not obj.file and not obj.external_url:
                messages.error(request, 'Please provide a file or an external URL.')
            else:
                obj.save()
                messages.success(request, 'Gallery media added successfully.')
                return redirect('admin_gallery')
    else:
        form = ContactGalleryForm()

    return render(request, 'store/admin/gallery.html', {
        'contact': contact,
        'media_items': media_items,
        'form': form,
    })


@login_required
@user_passes_test(is_admin)
def admin_delete_gallery_media_view(request, media_id):
    media = get_object_or_404(ContactGalleryMedia, id=media_id)
    media.delete()
    messages.success(request, 'Gallery media deleted successfully.')
    return redirect('admin_gallery')


@login_required
@user_passes_test(is_admin)
def admin_reviews_view(request):
    pending_reviews = Review.objects.filter(approved=False).select_related('user', 'order').order_by('-created_at')
    approved_reviews = Review.objects.filter(approved=True).select_related('user', 'order').order_by('-created_at')
    
    return render(request, 'store/admin/reviews.html', {
        'pending_reviews': pending_reviews,
        'approved_reviews': approved_reviews,
    })


@login_required
@user_passes_test(is_admin)
def admin_approve_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    review.approved = True
    review.save()
    messages.success(request, 'Review approved successfully.')
    return redirect('admin_reviews')


@login_required
@user_passes_test(is_admin)
def admin_reject_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    review.delete()
    messages.success(request, 'Review rejected and deleted.')
    return redirect('admin_reviews')


@login_required
@user_passes_test(is_admin)
def admin_orders_view(request):
    orders = Order.objects.all().order_by('-created_at')
    
    # Filter
    form = OrderFilterForm(request.GET)
    if form.is_valid():
        status = form.cleaned_data.get('status')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        
        if status:
            orders = orders.filter(status=status)
        
        if start_date:
            orders = orders.filter(created_at__gte=start_date)
        
        if end_date:
            orders = orders.filter(created_at__lte=end_date)
    
    return render(request, 'store/admin/orders.html', {
        'orders': orders,
        'form': form,
    })


# AJAX endpoint for admin orders filter/search
@login_required
@user_passes_test(is_admin)
@require_GET
def admin_orders_ajax_view(request):
    orders = Order.objects.all().order_by('-created_at')
    form = OrderFilterForm(request.GET)
    if form.is_valid():
        status = form.cleaned_data.get('status')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        if status:
            orders = orders.filter(status=status)
        if start_date:
            orders = orders.filter(created_at__gte=start_date)
        if end_date:
            orders = orders.filter(created_at__lte=end_date)
    search = request.GET.get('search', '').strip()
    if search:
        orders = orders.filter(
            Q(order_number__icontains=search) |
            Q(user__username__icontains=search)
        )
    html = render_to_string('store/admin/_orders_table.html', {'orders': orders}, request=request)
    return JsonResponse({'html': html})


@login_required
@user_passes_test(is_admin)
def admin_order_detail_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'send_message':
            message_body = (request.POST.get('message') or '').strip()
            subject = (request.POST.get('subject') or '').strip()
            recipient = getattr(order.user, 'email', '') or ''

            if not recipient:
                messages.error(request, 'Cannot send email: customer has no email address on file.')
                return redirect('admin_order_detail', order_id=order.id)

            if not message_body:
                messages.error(request, 'Please enter a message before sending.')
                return redirect('admin_order_detail', order_id=order.id)

            subject_line = subject or f'Update on Order {order.order_number}'
            from_email = settings.DEFAULT_FROM_EMAIL or 'noreply@aquafishstore.com'

            try:
                send_mail(subject_line, message_body, from_email, [recipient], fail_silently=False)
            except Exception:
                logging.getLogger(__name__).exception('Failed to send admin order email for order %s', order.id)
                messages.error(request, 'Failed to send email. Please try again later.')
            else:
                messages.success(request, 'Email sent to customer successfully.')

            return redirect('admin_order_detail', order_id=order.id)

        new_status = request.POST.get('status')
        if new_status and new_status in dict(Order.STATUS_CHOICES):
            order.status = new_status
            order.save()
            messages.success(request, f'Order status updated to {order.get_status_display()}.')
            return redirect('admin_order_detail', order_id=order.id)
    
    return render(request, 'store/admin/order_detail.html', {'order': order})


@login_required
@user_passes_test(is_admin)
def admin_users_view(request):
    users = CustomUser.objects.filter(role='customer').order_by('-is_favorite', '-created_at')
    return render(request, 'store/admin/users.html', {'users': users})


@login_required
@user_passes_test(is_admin)
@require_GET
def admin_users_ajax_view(request):
    search_term = request.GET.get('search', '').strip()
    users = CustomUser.objects.filter(role='customer')
    if search_term:
        users = users.filter(
            Q(username__icontains=search_term) |
            Q(email__icontains=search_term) |
            Q(phone_number__icontains=search_term)
        )
    users = users.order_by('-is_favorite', '-created_at')
    html = render_to_string('store/admin/_users_table.html', {'users': users}, request=request)
    return JsonResponse({'html': html})


@login_required
@user_passes_test(is_admin)
def toggle_favorite_user_view(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id, role='customer')
    user.is_favorite = not user.is_favorite
    user.save()
    status = 'added to' if user.is_favorite else 'removed from'
    messages.success(request, f'User {user.username} has been {status} favorites.')
    return redirect('admin_users')


@login_required
@user_passes_test(is_admin)
def block_user_view(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id, role='customer')
    user.is_blocked = True
    user.save()
    # Terminate any active sessions for this user so they are immediately logged out
    removed = terminate_user_sessions(user)
    if removed:
        messages.success(request, f'User {user.username} has been blocked and {removed} active session(s) terminated.')
    else:
        messages.success(request, f'User {user.username} has been blocked.')
    return redirect('admin_users')


@login_required
@user_passes_test(is_admin)
def unblock_user_view(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id, role='customer')
    user.is_blocked = False
    user.save()
    messages.success(request, f'User {user.username} has been unblocked.')
    return redirect('admin_users')


@login_required
@user_passes_test(is_admin)
def export_orders_excel_view(request):
    orders = Order.objects.all().order_by('-created_at')
    
    # Filter
    status = request.GET.get('status')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if status:
        orders = orders.filter(status=status)
    if start_date:
        orders = orders.filter(created_at__gte=start_date)
    if end_date:
        orders = orders.filter(created_at__lte=end_date)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Orders Summary
    ws1 = wb.active
    ws1.title = "Orders Summary"
    
    # Orders Summary Headers
    summary_headers = [
        'Order Number', 'Customer Username', 'Customer Email', 'Customer Phone', 
        'Total Amount', 'Order Status', 'Payment Method', 'Payment Status', 'Transaction ID',
        'Shipping Address', 'Contact Phone', 'Total Items', 'Order Created', 'Last Updated'
    ]
    ws1.append(summary_headers)
    
    # Style header
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Orders Summary Data
    for order in orders:
        total_items = order.items.aggregate(total=models.Sum('quantity'))['total'] or 0
        ws1.append([
            order.order_number,
            order.user.username,
            order.user.email,
            order.user.phone_number if order.user.phone_number else 'N/A',
            float(order.total_amount),
            order.get_status_display(),
            order.get_payment_method_display(),
            order.get_payment_status_display(),
            order.transaction_id if order.transaction_id else 'N/A',
            order.shipping_address,
            order.phone_number,
            total_items,
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
        ])
    
    # Auto-adjust column widths for summary sheet
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Detailed Order Items
    ws2 = wb.create_sheet(title="Order Items Detail")
    
    # Order Items Headers
    items_headers = [
        'Order Number', 'Customer Name', 'Fish Name', 'Fish Breed', 'Fish Category',
        'Quantity', 'Unit Price', 'Item Total', 'Order Status', 'Order Date'
    ]
    ws2.append(items_headers)
    
    # Style items header
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Order Items Data
    for order in orders:
        for item in order.items.all():
            ws2.append([
                order.order_number,
                order.user.username,
                item.fish.name,
                item.fish.breed.name,
                item.fish.category.name,
                item.quantity,
                float(item.price),
                float(item.get_total()),
                order.get_status_display(),
                order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ])
    
    # Auto-adjust column widths for items sheet
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 3: Customer Information
    ws3 = wb.create_sheet(title="Customer Details")
    
    # Get unique customers from filtered orders
    customer_ids = orders.values_list('user_id', flat=True).distinct()
    customers = CustomUser.objects.filter(id__in=customer_ids)
    
    # Customer Headers
    customer_headers = [
        'Username', 'Email', 'Phone Number', 'Date Joined', 
        'Email Verified', 'Total Orders', 'Total Spent'
    ]
    ws3.append(customer_headers)
    
    # Style customer header
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Customer Data
    for customer in customers:
        customer_orders = orders.filter(user=customer)
        total_orders = customer_orders.count()
        total_spent = customer_orders.aggregate(total=models.Sum('total_amount'))['total'] or 0
        
        ws3.append([
            customer.username,
            customer.email,
            customer.phone_number if customer.phone_number else 'N/A',
            customer.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'Yes' if customer.email_verified else 'No',
            total_orders,
            float(total_spent),
        ])
    
    # Auto-adjust column widths for customer sheet
    for column in ws3.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws3.column_dimensions[column_letter].width = adjusted_width
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="orders_detailed_report.xlsx"'
    wb.save(response)
    return response


# Coupon Management Views
@login_required
@user_passes_test(is_admin)
def admin_coupons_view(request):
    filter_type = request.GET.get('filter', 'all_coupons')
    
    if filter_type == 'all_users':
        coupons = Coupon.objects.filter(coupon_type='all').order_by('-created_at')
    elif filter_type == 'favorites':
        coupons = Coupon.objects.filter(coupon_type='favorites').order_by('-created_at')
    elif filter_type == 'normal':
        coupons = Coupon.objects.filter(coupon_type='normal').order_by('-created_at')
    elif filter_type == 'active':
        coupons = Coupon.objects.filter(is_active=True).order_by('-created_at')
    elif filter_type == 'inactive':
        coupons = Coupon.objects.filter(is_active=False).order_by('-created_at')
    else:
        coupons = Coupon.objects.all().order_by('-created_at')
    
    # Get counts for badges
    all_users_count = Coupon.objects.filter(coupon_type='all').count()
    favorites_count = Coupon.objects.filter(coupon_type='favorites').count()
    normal_count = Coupon.objects.filter(coupon_type='normal').count()
    active_count = Coupon.objects.filter(is_active=True).count()
    inactive_count = Coupon.objects.filter(is_active=False).count()
    
    context = {
        'coupons': coupons,
        'filter_type': filter_type,
        'all_users_count': all_users_count,
        'favorites_count': favorites_count,
        'normal_count': normal_count,
        'active_count': active_count,
        'inactive_count': inactive_count,
    }
    return render(request, 'store/admin/coupons.html', context)


@login_required
@user_passes_test(is_admin)
def admin_add_coupon_view(request):
    if request.method == 'POST':
        form = CouponForm(request.POST)
        if form.is_valid():
            coupon = form.save(commit=False)
            coupon.created_by = request.user
            coupon.save()
            messages.success(request, f'Coupon "{coupon.code}" created successfully!')
            return redirect('admin_coupons')
    else:
        form = CouponForm()
    return render(request, 'store/admin/add_coupon.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def admin_edit_coupon_view(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    if request.method == 'POST':
        form = CouponForm(request.POST, instance=coupon)
        if form.is_valid():
            form.save()
            messages.success(request, f'Coupon "{coupon.code}" updated successfully!')
            return redirect('admin_coupons')
    else:
        form = CouponForm(instance=coupon)
    return render(request, 'store/admin/add_coupon.html', {'form': form, 'is_edit': True, 'coupon': coupon})


@login_required
@user_passes_test(is_admin)
def admin_delete_coupon_view(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    code = coupon.code
    coupon.delete()
    messages.success(request, f'Coupon "{code}" deleted successfully!')
    return redirect('admin_coupons')


@login_required
@user_passes_test(is_admin)
def admin_toggle_coupon_view(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    coupon.is_active = not coupon.is_active
    coupon.save()
    status = 'activated' if coupon.is_active else 'deactivated'
    messages.success(request, f'Coupon "{coupon.code}" has been {status}!')
    return redirect('admin_coupons')


# Profile Views
@login_required
def profile_view(request):
    return render(request, 'store/profile.html', {'user': request.user})


@login_required
def edit_profile_view(request):
    if request.method == 'POST':
        form = ProfileEditForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('profile')
    else:
        form = ProfileEditForm(instance=request.user)
    
    return render(request, 'store/edit_profile.html', {'form': form})


@login_required
def change_password_view(request):
    # Only allow users who have verified email to change password
    if not request.user.email_verified:
        messages.error(request, 'Please verify your email before changing password.')
        return redirect('profile')

    if request.method == 'POST':
        form = ChangePasswordForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Keep user logged in after password change
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been changed successfully.')
            return redirect('profile')
    else:
        form = ChangePasswordForm(request.user)

    return render(request, 'store/change_password.html', {'form': form})


from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
import razorpay

def start_payment(request):
    amount = 50000  # amount in paise (₹500)

    client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

    # Create order
    payment = client.order.create({
        "amount": amount,
        "currency": "INR",
        "payment_capture": "1",  # auto capture
    })

    # Pass order to frontend
    context = {
        "order_id": payment["id"],
        "amount": amount,
        "razorpay_key": settings.RAZORPAY_KEY_ID,
    }
    return render(request, "payment.html", context)